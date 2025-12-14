from __future__ import annotations

import csv
import os
from collections.abc import Iterable, Iterator, Sequence
from contextlib import contextmanager
from datetime import UTC, datetime
from pathlib import Path

try:  # pragma: no cover - optional dependency for Excel datasets
    from openpyxl import load_workbook
except ImportError:  # pragma: no cover - handled during hydration
    load_workbook = None

from sqlalchemy import Select, create_engine, delete, func, select
from sqlalchemy.engine import Engine
from sqlalchemy.orm import Session, joinedload, sessionmaker

from .migrations import run_migrations
from .schema import (
    AuditStatus,
    Base,
    BundleAudit,
    ClusterMembership,
    ColumnPreference,
    ColumnPreferenceChange,
    DataFile,
    EmbeddingVector,
    FileType,
    IngestionAudit,
    IngestionStatus,
    MetricType,
    PerformanceMetric,
    PreferenceMirror,
    QueryDefinition,
    QueryRecord,
    QuerySheetLink,
    QuerySheetRole,
    SheetMetric,
    SheetMetricType,
    SheetSource,
    SheetStatus,
    SheetVisibilityState,
    SimilarityCluster,
    SourceBundle,
)

DEFAULT_SQLITE_URL = "sqlite:///data/metadata.db"


def _resolve_sqlite_url(url: str | None = None) -> str:
    resolved = url or os.getenv("SQLITE_URL", DEFAULT_SQLITE_URL)
    if resolved.startswith("sqlite:///"):
        db_path = Path(resolved.replace("sqlite:///", "", 1))
        db_path.parent.mkdir(parents=True, exist_ok=True)
    return resolved


def build_engine(url: str | None = None) -> Engine:
    resolved = _resolve_sqlite_url(url)
    connect_args = {"check_same_thread": False} if resolved.startswith("sqlite") else {}
    return create_engine(resolved, future=True, connect_args=connect_args)


def create_session_factory(engine: Engine) -> sessionmaker[Session]:
    return sessionmaker(bind=engine, class_=Session, expire_on_commit=False, future=True)


def init_database(engine: Engine) -> None:
    Base.metadata.create_all(engine)
    run_migrations(engine)


@contextmanager
def session_scope(session_factory: sessionmaker[Session]) -> Iterator[Session]:
    session = session_factory()
    try:
        yield session
        session.commit()
    except Exception:
        session.rollback()
        raise
    finally:
        session.close()


class MetadataRepository:
    """Data access helpers for SQLite-backed metadata and analytics."""

    def __init__(self, session: Session) -> None:
        self.session = session
        self._row_cache: dict[tuple[str, str | None], list[dict[str, object]]] = {}
        self._sheet_cache: dict[str, SheetSource | None] = {}

    # Dataset helpers -----------------------------------------------------
    def list_data_files(self) -> Sequence[DataFile]:
        stmt: Select[tuple[DataFile]] = select(DataFile)
        return self.session.execute(stmt).scalars().all()

    def get_data_file(self, data_file_id: str) -> DataFile | None:
        return self.session.get(DataFile, data_file_id)

    def get_data_file_by_hash(self, file_hash: str) -> DataFile | None:
        stmt: Select[tuple[DataFile]] = select(DataFile).where(DataFile.file_hash == file_hash)
        return self.session.execute(stmt).scalars().first()

    def create_data_file(
        self,
        *,
        display_name: str,
        original_path: str,
        file_hash: str,
        file_type: FileType,
        delimiter: str | None,
        sheet_name: str | None,
        selected_columns: Sequence[str],
        status: IngestionStatus = IngestionStatus.PENDING,
    ) -> DataFile:
        data_file = DataFile(
            display_name=display_name,
            original_path=original_path,
            file_hash=file_hash,
            file_type=file_type,
            delimiter=delimiter,
            sheet_name=sheet_name,
            selected_columns=list(selected_columns),
            ingestion_status=status,
        )
        self.session.add(data_file)
        return data_file

    def update_data_file_status(
        self,
        data_file: DataFile,
        *,
        status: IngestionStatus,
        row_count: int | None = None,
        error_summary: str | None = None,
        processed_at: datetime | None = None,
    ) -> DataFile:
        data_file.ingestion_status = status
        if row_count is not None:
            data_file.row_count = row_count
        data_file.error_summary = error_summary
        if processed_at is not None:
            data_file.processed_at = processed_at
        return data_file

    # Source bundle helpers -------------------------------------------
    def list_source_bundles(self) -> Sequence[SourceBundle]:
        stmt: Select[tuple[SourceBundle]] = select(SourceBundle).order_by(
            SourceBundle.created_at.desc()
        )
        return self.session.execute(stmt).scalars().all()

    def get_source_bundle(self, bundle_id: str) -> SourceBundle | None:
        return self.session.get(SourceBundle, bundle_id)

    def get_source_bundle_by_hash(self, file_hash: str) -> SourceBundle | None:
        stmt: Select[tuple[SourceBundle]] = select(SourceBundle).where(
            SourceBundle.file_hash == file_hash
        )
        return self.session.execute(stmt).scalars().first()

    def create_source_bundle(
        self,
        *,
        display_name: str,
        original_path: str,
        file_hash: str,
        file_type: FileType,
        delimiter: str | None,
        refresh_cadence: str | None,
        owner_user_id: str | None = None,
        ingestion_status: IngestionStatus = IngestionStatus.PENDING,
    ) -> SourceBundle:
        bundle = SourceBundle(
            display_name=display_name,
            original_path=original_path,
            file_hash=file_hash,
            file_type=file_type,
            delimiter=delimiter,
            refresh_cadence=refresh_cadence,
            ingestion_status=ingestion_status,
            owner_user_id=owner_user_id,
        )
        self.session.add(bundle)
        return bundle

    def update_source_bundle(
        self,
        bundle: SourceBundle,
        *,
        ingestion_status: IngestionStatus | None = None,
        sheet_count: int | None = None,
        refresh_cadence: str | None = None,
        original_path: str | None = None,
    ) -> SourceBundle:
        if ingestion_status is not None:
            bundle.ingestion_status = ingestion_status
        if sheet_count is not None:
            bundle.sheet_count = sheet_count
        if refresh_cadence is not None:
            bundle.refresh_cadence = refresh_cadence
        if original_path is not None:
            bundle.original_path = original_path
        bundle.updated_at = datetime.now(UTC)
        return bundle

    # Sheet source helpers --------------------------------------------
    def list_sheet_sources(
        self,
        *,
        bundle_id: str | None = None,
        statuses: Sequence[SheetStatus] | None = None,
    ) -> Sequence[SheetSource]:
        stmt: Select[tuple[SheetSource]] = select(SheetSource)
        if bundle_id:
            stmt = stmt.where(SheetSource.bundle_id == bundle_id)
        if statuses:
            stmt = stmt.where(SheetSource.status.in_(tuple(statuses)))
        stmt = stmt.order_by(SheetSource.position_index.asc())
        return self.session.execute(stmt).scalars().all()

    def get_sheet_source(self, sheet_id: str) -> SheetSource | None:
        return self.session.get(SheetSource, sheet_id)

    def create_sheet_source(
        self,
        *,
        bundle: SourceBundle,
        sheet_name: str,
        display_label: str,
        visibility_state: SheetVisibilityState,
        status: SheetStatus,
        row_count: int,
        column_schema: Sequence[dict[str, object]],
        position_index: int,
        checksum: str | None,
        description: str | None = None,
        tags: Sequence[str] | None = None,
        last_refreshed_at: datetime | None = None,
    ) -> SheetSource:
        sheet = SheetSource(
            bundle=bundle,
            sheet_name=sheet_name,
            display_label=display_label,
            visibility_state=visibility_state,
            status=status,
            row_count=row_count,
            column_schema=list(column_schema),
            position_index=position_index,
            checksum=checksum,
            description=description,
            tags=list(tags) if tags else None,
            last_refreshed_at=last_refreshed_at,
        )
        self.session.add(sheet)
        return sheet

    def update_sheet_source(
        self,
        sheet: SheetSource,
        *,
        sheet_name: str | None = None,
        display_label: str | None = None,
        visibility_state: SheetVisibilityState | None = None,
        status: SheetStatus | None = None,
        row_count: int | None = None,
        column_schema: Sequence[dict[str, object]] | None = None,
        checksum: str | None = None,
        description: str | None = None,
        tags: Sequence[str] | None = None,
        last_refreshed_at: datetime | None = None,
        position_index: int | None = None,
    ) -> SheetSource:
        if sheet_name is not None:
            sheet.sheet_name = sheet_name
        if display_label is not None:
            sheet.display_label = display_label
        if visibility_state is not None:
            sheet.visibility_state = visibility_state
        if status is not None:
            sheet.status = status
        if row_count is not None:
            sheet.row_count = row_count
        if column_schema is not None:
            sheet.column_schema = list(column_schema)
        if checksum is not None:
            sheet.checksum = checksum
        if description is not None:
            sheet.description = description
        if tags is not None:
            sheet.tags = list(tags)
        if last_refreshed_at is not None:
            sheet.last_refreshed_at = last_refreshed_at
        if position_index is not None:
            sheet.position_index = position_index
        return sheet

    def fetch_search_candidates(
        self,
        *,
        dataset_ids: Sequence[str] | None = None,
        column_names: Sequence[str] | None = None,
        max_records: int | None = 5000,
    ) -> Sequence[QueryRecord]:
        stmt: Select[tuple[QueryRecord]] = (
            select(QueryRecord)
            .options(
                joinedload(QueryRecord.data_file),
                joinedload(QueryRecord.sheet),
            )
            .order_by(QueryRecord.created_at.desc())
        )
        if dataset_ids:
            stmt = stmt.where(QueryRecord.data_file_id.in_(tuple(dataset_ids)))
        if column_names:
            stmt = stmt.where(QueryRecord.column_name.in_(tuple(column_names)))
        if max_records is not None:
            stmt = stmt.limit(max(1, max_records))
        return self.session.execute(stmt).scalars().all()

    # Query records -------------------------------------------------------
    def replace_query_records(
        self,
        data_file: DataFile,
        records: Iterable[QueryRecord],
    ) -> None:
        existing = select(QueryRecord).where(QueryRecord.data_file_id == data_file.id)
        for record in self.session.execute(existing).scalars():
            self.session.delete(record)
        for record in records:
            self.session.add(record)

    def create_query_record(
        self,
        *,
        data_file_id: str,
        column_name: str,
        row_index: int,
        text: str,
        original_text: str,
        tags: Sequence[str] | None = None,
        sheet_id: str | None = None,
    ) -> QueryRecord:
        record = QueryRecord(
            data_file_id=data_file_id,
            column_name=column_name,
            row_index=row_index,
            text=text,
            original_text=original_text,
            tags=list(tags) if tags else None,
            sheet_id=sheet_id,
        )
        self.session.add(record)
        return record

    def delete_query_records_for_sheet(self, sheet_id: str) -> None:
        stmt: Select[tuple[QueryRecord]] = select(QueryRecord).where(
            QueryRecord.sheet_id == sheet_id
        )
        for record in self.session.execute(stmt).scalars():
            self.session.delete(record)

    # Sheet lifecycle audits ------------------------------------------
    def create_bundle_audit(
        self,
        *,
        bundle: SourceBundle,
        status: AuditStatus,
        started_at: datetime,
        completed_at: datetime | None,
        sheet_summary: dict[str, int] | None = None,
        hidden_sheets_enabled: Sequence[str] | None = None,
        initiated_by: str | None = None,
    ) -> BundleAudit:
        audit = BundleAudit(
            bundle=bundle,
            status=status,
            started_at=started_at,
            completed_at=completed_at,
            sheet_summary=dict(sheet_summary) if sheet_summary else None,
            hidden_sheets_enabled=list(hidden_sheets_enabled) if hidden_sheets_enabled else None,
            initiated_by=initiated_by,
        )
        self.session.add(audit)
        return audit

    def get_latest_bundle_audit(self, bundle_id: str) -> BundleAudit | None:
        stmt = (
            select(BundleAudit)
            .where(BundleAudit.bundle_id == bundle_id)
            .order_by(BundleAudit.started_at.desc())
        )
        return self.session.execute(stmt).scalars().first()

    # Sheet metrics ----------------------------------------------------
    def record_sheet_metric(
        self,
        *,
        sheet: SheetSource,
        metric_type: SheetMetricType,
        p50: float | None,
        p95: float | None,
        recorded_at: datetime | None = None,
    ) -> SheetMetric:
        metric = SheetMetric(
            sheet=sheet,
            metric_type=metric_type,
            p50=p50,
            p95=p95,
            recorded_at=recorded_at or datetime.now(UTC),
        )
        self.session.add(metric)
        return metric

    # Query definitions ------------------------------------------------
    def list_query_definitions(self) -> Sequence[QueryDefinition]:
        stmt: Select[tuple[QueryDefinition]] = select(QueryDefinition).order_by(
            QueryDefinition.created_at.desc()
        )
        return self.session.execute(stmt).scalars().all()

    def get_query_definition(self, query_id: str) -> QueryDefinition | None:
        return self.session.get(QueryDefinition, query_id)

    def create_query_definition(
        self,
        *,
        name: str,
        definition: dict[str, object],
        description: str | None = None,
        validation_checksum: str | None = None,
    ) -> QueryDefinition:
        query = QueryDefinition(
            name=name,
            description=description,
            definition=dict(definition),
            validation_checksum=validation_checksum,
        )
        self.session.add(query)
        return query

    def update_query_definition(
        self,
        query: QueryDefinition,
        *,
        description: str | None = None,
        definition: dict[str, object] | None = None,
        validation_checksum: str | None = None,
    ) -> QueryDefinition:
        if description is not None:
            query.description = description
        if definition is not None:
            query.definition = dict(definition)
        if validation_checksum is not None:
            query.validation_checksum = validation_checksum
        query.updated_at = datetime.now(UTC)
        return query

    def create_query_with_links(
        self,
        *,
        name: str,
        definition: dict[str, object],
        description: str | None,
        sheet_links: Sequence[tuple[str, QuerySheetRole, Sequence[str] | None, datetime | None]],
        validation_checksum: str | None = None,
    ) -> QueryDefinition:
        query = self.create_query_definition(
            name=name,
            description=description,
            definition=definition,
            validation_checksum=validation_checksum,
        )
        self.set_query_sheet_links(query=query, links=sheet_links)
        return query

    def update_query_with_links(
        self,
        *,
        query: QueryDefinition,
        definition: dict[str, object] | None,
        description: str | None,
        sheet_links: Sequence[tuple[str, QuerySheetRole, Sequence[str] | None, datetime | None]],
        validation_checksum: str | None = None,
    ) -> QueryDefinition:
        updated = self.update_query_definition(
            query,
            description=description,
            definition=definition,
            validation_checksum=validation_checksum,
        )
        self.set_query_sheet_links(query=updated, links=sheet_links)
        return updated

    def set_query_sheet_links(
        self,
        *,
        query: QueryDefinition,
        links: Sequence[tuple[str, QuerySheetRole, Sequence[str] | None, datetime | None]],
    ) -> None:
        self.session.execute(delete(QuerySheetLink).where(QuerySheetLink.query_id == query.id))
        for sheet_id, role, join_keys, last_validated_at in links:
            link = QuerySheetLink(
                query_id=query.id,
                sheet_id=sheet_id,
                role=role,
                join_keys=list(join_keys) if join_keys else None,
                last_validated_at=last_validated_at,
            )
            self.session.add(link)

    def list_query_links_for_sheet(self, sheet_id: str) -> Sequence[QuerySheetLink]:
        stmt: Select[tuple[QuerySheetLink]] = select(QuerySheetLink).where(
            QuerySheetLink.sheet_id == sheet_id
        )
        return self.session.execute(stmt).scalars().all()

    def list_query_definitions_with_links(self) -> Sequence[QueryDefinition]:
        stmt: Select[tuple[QueryDefinition]] = (
            select(QueryDefinition)
            .options(joinedload(QueryDefinition.sheet_links))
            .order_by(QueryDefinition.created_at.desc())
        )
        return self.session.execute(stmt).scalars().all()

    # Embeddings ----------------------------------------------------------
    def upsert_embedding(
        self,
        *,
        record_id: str,
        model_name: str,
        model_version: str,
        vector_path: str,
        embedding_dim: int,
    ) -> EmbeddingVector:
        stmt = select(EmbeddingVector).where(EmbeddingVector.query_record_id == record_id)
        embedding = self.session.execute(stmt).scalars().first()
        if embedding is None:
            embedding = EmbeddingVector(
                query_record_id=record_id,
                model_name=model_name,
                model_version=model_version,
                vector_path=vector_path,
                embedding_dim=embedding_dim,
            )
            self.session.add(embedding)
        else:
            embedding.model_name = model_name
            embedding.model_version = model_version
            embedding.vector_path = vector_path
            embedding.embedding_dim = embedding_dim
        return embedding

    def get_sheet_embedding_counts(self, sheet_ids: Sequence[str] | None = None) -> dict[str, int]:
        stmt = (
            select(QueryRecord.sheet_id, func.count(EmbeddingVector.id))
            .join(EmbeddingVector, EmbeddingVector.query_record_id == QueryRecord.id)
            .group_by(QueryRecord.sheet_id)
        )
        if sheet_ids:
            stmt = stmt.where(QueryRecord.sheet_id.in_(tuple(sheet_ids)))
        rows = self.session.execute(stmt).all()
        counts: dict[str, int] = {}
        for sheet_id, count in rows:
            if sheet_id is None:
                continue
            counts[sheet_id] = int(count or 0)
        if sheet_ids:
            for sheet_id in sheet_ids:
                counts.setdefault(sheet_id, 0)
        return counts

    # Clusters ------------------------------------------------------------
    def save_similarity_clusters(
        self,
        *,
        clusters: Iterable[SimilarityCluster],
        memberships: Iterable[ClusterMembership],
        clear_existing: bool = True,
    ) -> None:
        if clear_existing:
            self.session.execute(delete(ClusterMembership))
            self.session.execute(delete(SimilarityCluster))
        for cluster in clusters:
            self.session.merge(cluster)
        for membership in memberships:
            self.session.merge(membership)

    def list_similarity_clusters(self) -> Sequence[SimilarityCluster]:
        stmt: Select[tuple[SimilarityCluster]] = select(SimilarityCluster)
        return self.session.execute(stmt).scalars().all()

    # Audits --------------------------------------------------------------
    def create_audit(
        self,
        *,
        data_file_id: str,
        status: AuditStatus,
        processed_rows: int,
        skipped_rows: int,
        error_log_path: str | None = None,
        started_at: datetime | None = None,
        completed_at: datetime | None = None,
    ) -> IngestionAudit:
        audit = IngestionAudit(
            data_file_id=data_file_id,
            status=status,
            processed_rows=processed_rows,
            skipped_rows=skipped_rows,
            error_log_path=error_log_path,
            started_at=started_at or datetime.now(UTC),
            completed_at=completed_at,
        )
        self.session.add(audit)
        return audit

    def get_latest_audit(self, data_file_id: str) -> IngestionAudit | None:
        stmt = (
            select(IngestionAudit)
            .where(IngestionAudit.data_file_id == data_file_id)
            .order_by(IngestionAudit.started_at.desc())
        )
        return self.session.execute(stmt).scalars().first()

    # Column preferences ------------------------------------------------
    def get_column_preference(
        self,
        *,
        data_file_id: str,
        user_id: str | None,
        include_inactive: bool = False,
    ) -> ColumnPreference | None:
        stmt = select(ColumnPreference).where(ColumnPreference.data_file_id == data_file_id)
        if user_id is None:
            stmt = stmt.where(ColumnPreference.user_id.is_(None))
        else:
            stmt = stmt.where(ColumnPreference.user_id == user_id)
        if not include_inactive:
            stmt = stmt.where(ColumnPreference.is_active.is_(True))
        stmt = stmt.order_by(ColumnPreference.updated_at.desc())
        return self.session.execute(stmt).scalars().first()

    def save_column_preference(
        self,
        *,
        data_file_id: str,
        user_id: str | None,
        selected_columns: Sequence[dict[str, object]],
        max_columns: int,
        allowed_columns: set[str] | None = None,
        actor_user_id: str | None = None,
    ) -> ColumnPreference:
        normalized = self._normalize_preference_columns(
            selected_columns,
            data_file_id=data_file_id,
            max_columns=max_columns,
            allowed_columns=allowed_columns,
        )
        record = self.get_column_preference(
            data_file_id=data_file_id,
            user_id=user_id,
            include_inactive=True,
        )
        if record is None:
            record = ColumnPreference(
                data_file_id=data_file_id,
                user_id=user_id,
                selected_columns=normalized,
                max_columns=max_columns,
                is_active=True,
            )
            self.session.add(record)
        else:
            record.selected_columns = normalized
            record.max_columns = max_columns
            record.is_active = True
            record.updated_at = datetime.now(UTC)
        self.session.flush()
        self._record_preference_change(
            preference=record,
            user_id=actor_user_id or user_id,
        )
        return record

    def reset_column_preference(
        self,
        *,
        data_file_id: str,
        user_id: str | None,
        actor_user_id: str | None = None,
    ) -> None:
        record = self.get_column_preference(
            data_file_id=data_file_id,
            user_id=user_id,
            include_inactive=True,
        )
        if record is None:
            return
        record.is_active = False
        record.selected_columns = []
        record.updated_at = datetime.now(UTC)
        self.session.flush()
        self._record_preference_change(
            preference=record,
            user_id=actor_user_id or user_id,
        )

    def get_preference_mirror(
        self,
        *,
        data_file_id: str,
        device_id: str | None,
    ) -> PreferenceMirror | None:
        stmt = select(PreferenceMirror).where(PreferenceMirror.data_file_id == data_file_id)
        if device_id is None:
            stmt = stmt.where(PreferenceMirror.device_id.is_(None))
        else:
            stmt = stmt.where(PreferenceMirror.device_id == device_id)
        stmt = stmt.order_by(PreferenceMirror.updated_at.desc())
        return self.session.execute(stmt).scalars().first()

    def upsert_preference_mirror(
        self,
        *,
        data_file_id: str,
        device_id: str | None,
        selected_columns: Sequence[dict[str, object]],
        max_columns: int,
        version: int = 0,
        source: str | None = None,
    ) -> PreferenceMirror:
        sanitized_max = max(max_columns, 1)
        normalized = self._normalize_preference_columns(
            selected_columns,
            data_file_id=data_file_id,
            max_columns=sanitized_max,
            allowed_columns=None,
        )

        record = self.get_preference_mirror(data_file_id=data_file_id, device_id=device_id)
        if record is None:
            record = PreferenceMirror(
                data_file_id=data_file_id,
                device_id=device_id,
                selected_columns=normalized,
                max_columns=sanitized_max,
                version=version,
                source=source,
                updated_at=datetime.now(UTC),
            )
            self.session.add(record)
        else:
            record.selected_columns = normalized
            record.max_columns = sanitized_max
            record.version = version
            record.source = source or record.source
            record.updated_at = datetime.now(UTC)

        self.session.flush()
        return record

    def list_displayable_column_catalog(self, data_file_id: str) -> list[dict[str, object]]:
        data_file = self.get_data_file(data_file_id)
        if data_file is None:
            return []

        rows = self._load_dataset_rows(data_file, sheet_name=None)
        sample_limit = 500
        samples: dict[str, list[object | None]] = {}

        for row in rows:
            for column, value in row.items():
                if not column:
                    continue
                bucket = samples.setdefault(column, [])
                if len(bucket) < sample_limit:
                    bucket.append(value)

        if not samples:
            for column in data_file.selected_columns or []:
                if column:
                    samples.setdefault(column, [])

        timestamp = data_file.processed_at or data_file.ingested_at
        catalog: list[dict[str, object]] = []
        for column, values in sorted(samples.items(), key=lambda item: item[0].lower()):
            catalog.append(
                {
                    "column_name": column,
                    "column_label": column,
                    "data_type": self._infer_value_type(values),
                    "is_available": True,
                    "last_seen_at": timestamp,
                }
            )
        return catalog

    def list_column_preference_changes(
        self, preference_id: str
    ) -> Sequence[ColumnPreferenceChange]:
        stmt = (
            select(ColumnPreferenceChange)
            .where(ColumnPreferenceChange.preference_id == preference_id)
            .order_by(ColumnPreferenceChange.changed_at.desc())
        )
        return self.session.execute(stmt).scalars().all()

    def _normalize_preference_columns(
        self,
        selections: Sequence[dict[str, object]],
        *,
        data_file_id: str,
        max_columns: int,
        allowed_columns: set[str] | None,
    ) -> list[dict[str, object]]:
        if max_columns < 1:
            raise ValueError("max_columns must be at least 1.")

        normalized: list[dict[str, object]] = []
        seen: set[str] = set()
        sorted_entries = sorted(selections, key=lambda entry: int(entry.get("position", 0)))
        for entry in sorted_entries:
            name = str(entry.get("column_name", "")).strip()
            if not name:
                continue
            if allowed_columns is not None and name not in allowed_columns:
                raise ValueError(f"Unknown columns for dataset {data_file_id}: {name}.")
            if name in seen:
                raise ValueError(f"Duplicate columns are not allowed: {name}.")
            seen.add(name)
            label = str(entry.get("display_label", name)).strip() or name
            normalized.append(
                {
                    "column_name": name,
                    "display_label": label,
                    "position": len(normalized),
                }
            )

        if len(normalized) > max_columns:
            raise ValueError(f"Selection exceeds the maximum allowed columns ({max_columns}).")
        return normalized

    def _record_preference_change(
        self,
        *,
        preference: ColumnPreference,
        user_id: str | None,
        dataset_display_name: str | None = None,
    ) -> ColumnPreferenceChange:
        actor = user_id or "system"
        display_name = dataset_display_name
        if not display_name:
            data_file = preference.data_file or self.get_data_file(preference.data_file_id)
            display_name = data_file.display_name if data_file else preference.data_file_id
        change = ColumnPreferenceChange(
            preference_id=preference.id,
            user_id=actor,
            dataset_display_name=display_name,
            selected_columns_snapshot=list(preference.selected_columns),
        )
        self.session.add(change)
        return change

    def get_row_values(self, record: QueryRecord) -> dict[str, object]:
        data_file = record.data_file or self.get_data_file(record.data_file_id)
        if data_file is None:
            return {}

        sheet_name: str | None = None
        if record.sheet_id:
            sheet_name = self._resolve_sheet_name(record.sheet_id)

        rows = self._load_dataset_rows(data_file, sheet_name=sheet_name)
        if record.row_index < 0 or record.row_index >= len(rows):
            return {}
        return rows[record.row_index]

    # Performance metrics -------------------------------------------------
    def record_performance_metric(
        self,
        *,
        metric_type: MetricType,
        data_file_id: str | None,
        cluster_id: str | None,
        benchmark_run_id: str | None,
        p50_ms: float,
        p95_ms: float,
        records_per_second: float | None,
    ) -> PerformanceMetric:
        metric = PerformanceMetric(
            metric_type=metric_type,
            data_file_id=data_file_id,
            cluster_id=cluster_id,
            benchmark_run_id=benchmark_run_id,
            p50_ms=p50_ms,
            p95_ms=p95_ms,
            records_per_second=records_per_second,
        )
        self.session.add(metric)
        return metric

    def _resolve_sheet_name(self, sheet_id: str) -> str | None:
        if sheet_id in self._sheet_cache:
            sheet = self._sheet_cache[sheet_id]
            return sheet.sheet_name if sheet else None
        sheet = self.session.get(SheetSource, sheet_id)
        self._sheet_cache[sheet_id] = sheet
        if sheet is None:
            return None
        return sheet.sheet_name

    def _load_dataset_rows(
        self, data_file: DataFile, *, sheet_name: str | None
    ) -> list[dict[str, object]]:
        sheet_key = sheet_name or "__default__"
        cache_key = (data_file.id, sheet_key)
        if cache_key in self._row_cache:
            return self._row_cache[cache_key]

        path = Path(data_file.original_path)
        if not path.exists():
            self._row_cache[cache_key] = []
            return []

        if data_file.file_type == FileType.CSV:
            delimiter = data_file.delimiter or ","
            rows = self._load_csv_rows(path=path, delimiter=delimiter)
        else:
            rows = self._load_excel_rows(path=path, sheet_name=sheet_name or data_file.sheet_name)

        self._row_cache[cache_key] = rows
        return rows

    def _load_csv_rows(self, *, path: Path, delimiter: str) -> list[dict[str, object]]:
        rows: list[dict[str, object]] = []
        with path.open("r", encoding="utf-8", newline="") as handle:
            reader = csv.DictReader(handle, delimiter=delimiter)
            for raw_row in reader:
                row_map: dict[str, object] = {}
                for key, value in raw_row.items():
                    if not key:
                        continue
                    row_map[key] = value
                rows.append(row_map)
        return rows

    def _load_excel_rows(self, *, path: Path, sheet_name: str | None) -> list[dict[str, object]]:
        if load_workbook is None:
            raise RuntimeError("Excel contextual column hydration requires the 'openpyxl' package.")
        workbook = load_workbook(path, read_only=True, data_only=True)
        try:
            target_sheet = sheet_name or workbook.sheetnames[0]
            if target_sheet not in workbook.sheetnames:
                target_sheet = workbook.sheetnames[0]
            sheet = workbook[target_sheet]
            iterator = sheet.iter_rows(values_only=True)
            try:
                headers_row = next(iterator)
            except StopIteration:
                return []
            columns = [str(value) if value is not None else "" for value in headers_row]
            rows: list[dict[str, object]] = []
            for values in iterator:
                row_map: dict[str, object] = {}
                for index, column in enumerate(columns):
                    if not column:
                        continue
                    cell_value = None
                    if values is not None and index < len(values):
                        cell_value = values[index]
                    row_map[column] = cell_value
                rows.append(row_map)
            return rows
        finally:
            workbook.close()

    def _infer_value_type(self, values: Sequence[object | None]) -> str:
        non_null = [value for value in values if value not in (None, "")]
        if not non_null:
            return "string"
        if all(isinstance(value, int | float) for value in non_null):
            return "number"
        if all(isinstance(value, bool) for value in non_null):
            return "boolean"
        return "string"
