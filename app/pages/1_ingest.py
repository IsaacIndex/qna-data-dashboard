from __future__ import annotations

import csv
import io
import sys
import tempfile
from collections.abc import Sequence
from dataclasses import dataclass
from datetime import UTC, datetime
from pathlib import Path

import streamlit as st
from sqlalchemy.orm import Session, sessionmaker
from streamlit.runtime.uploaded_file_manager import UploadedFile

ROOT_DIR = Path(__file__).resolve().parents[2]
if str(ROOT_DIR) not in sys.path:
    sys.path.insert(0, str(ROOT_DIR))

try:  # Optional Excel dependency
    from openpyxl import load_workbook
except ImportError:  # pragma: no cover - optional path
    load_workbook = None

from app.db.metadata import (  # noqa: E402
    MetadataRepository,
    build_engine,
    create_session_factory,
    init_database,
    session_scope,
)
from app.db.schema import SheetStatus  # noqa: E402
from app.services.chroma_client import (  # noqa: E402
    get_chroma_client,
    get_chroma_runtime_state,
)
from app.services.embeddings import EmbeddingService  # noqa: E402
from app.services.ingestion import (  # noqa: E402
    BundleIngestionOptions,
    HiddenSheetPolicy,
    IngestionService,
    aggregate_column_catalog,
    build_column_picker_options,
)
from app.services.preferences import persist_column_selection  # noqa: E402
from app.utils.caching import cache_resource  # noqa: E402
from app.utils.config import get_data_root  # noqa: E402
from app.utils.logging import get_logger, log_event, log_timing  # noqa: E402
from app.utils.session_state import (
    confirm_reset,
    ensure_session_defaults,
    request_reset,
)  # noqa: E402
from app.services.ingest_models import SourceFile  # noqa: E402
from app.services.ingest_storage import default_storage  # noqa: E402
from app.services.embedding_queue import default_queue  # noqa: E402
from app.utils.audit import record_audit  # noqa: E402

LOGGER = get_logger(__name__)


def _list_document_groups() -> list[str]:
    root = default_storage.storage_root
    groups = ["default"]
    if root.exists():
        for item in root.iterdir():
            if item.is_dir() and not item.name.startswith("_"):
                groups.append(item.name)
    return sorted(set(groups))


def _render_source_manager() -> tuple[str, list[SourceFile]]:
    st.subheader("Source management")
    groups = _list_document_groups()
    group = st.selectbox("Document group", options=groups, index=0, key="ingest_group_select")

    sources: list[SourceFile] = default_storage.list_sources(group)
    st.caption(f"{len(sources)} sources in '{group}'")
    if sources:
        st.dataframe(
            [
                {
                    "Name": source.version_label,
                    "Size (KB)": round(source.size_bytes / 1024, 1),
                    "Status": source.status.value,
                    "Added": source.added_at,
                    "Columns": ", ".join(source.extracted_columns),
                }
                for source in sources
            ],
            hide_index=True,
            width="stretch",
        )
    files = st.file_uploader(
        "Add sources (CSV/XLS/XLSX/Parquet)",
        type=["csv", "xls", "xlsx", "parquet"],
        accept_multiple_files=True,
        key="ingest_source_uploader",
    )
    if files:
        for uploaded in files:
            try:
                default_storage.save_upload(
                    group,
                    uploaded,
                    filename=uploaded.name,
                    mime_type=uploaded.type or "",
                    added_by=None,
                )
                record_audit("ui.upload", "success", user=None, details={"group": group, "file": uploaded.name})
            except Exception as exc:  # noqa: BLE001
                st.error(f"Upload failed for {uploaded.name}: {exc}")
        st.success("Uploads processed. Refreshing list...")
        st.rerun()

    selected_for_delete = st.multiselect(
        "Delete sources",
        options=[(source.id, source.version_label) for source in sources],
        format_func=lambda item: item[1],
        key="ingest_delete_select",
    )
    if selected_for_delete and st.button("Confirm delete selected", type="primary"):
        for source_id, _label in selected_for_delete:
            default_storage.delete_source(group, source_id)
            record_audit("ui.delete", "success", user=None, details={"group": group, "source_id": source_id})
        st.success("Selected sources deleted.")
        st.rerun()

    return group, sources


@dataclass
class _UploadSheet:
    id: str
    display_label: str
    status: SheetStatus
    column_schema: list[dict[str, object]]
    last_refreshed_at: datetime | None = None


@cache_resource
def _get_session_factory() -> sessionmaker[Session]:
    engine = build_engine()
    init_database(engine)
    return create_session_factory(engine)


def _data_root() -> Path:
    return get_data_root()


def _render_persistence_status() -> None:
    _ = get_chroma_client()
    state = get_chroma_runtime_state()
    persist_path = state.persist_directory.expanduser()
    if state.is_persistent:
        st.success(
            f"Chroma persistence active — embeddings stored in `{persist_path}`.",
            icon="💾",
        )
        return

    if state.prefers_persistent:
        reason = "persistent client unavailable"
        if state.last_error:
            reason = f"{reason}: {state.last_error}"
        st.warning(
            f"Embeddings are running in-memory ({reason}). "
            f"They reset when the app restarts. Target directory: `{persist_path}`.",
            icon="⚠️",
        )
    else:
        st.warning(
            "Embeddings persistence disabled via `QNA_USE_CHROMADB=0`; "
            f"vectors reset between sessions. Target directory: `{persist_path}`.",
            icon="⚠️",
        )


def _preview_rows(
    uploaded_file: UploadedFile, sample_rows: int = 5
) -> tuple[list[str], list[dict[str, str]]]:
    uploaded_file.seek(0)
    if uploaded_file.name.lower().endswith(".csv"):
        content = uploaded_file.read().decode("utf-8")
        uploaded_file.seek(0)
        reader = csv.DictReader(io.StringIO(content))
        rows = []
        for index, row in enumerate(reader):
            rows.append(row)
            if index + 1 >= sample_rows:
                break
        return reader.fieldnames or [], rows

    if load_workbook is None:
        uploaded_file.seek(0)
        raise RuntimeError("Excel preview requires the 'openpyxl' dependency.")

    workbook = load_workbook(uploaded_file, read_only=True)
    sheet = workbook.active
    rows_iter = sheet.iter_rows(values_only=True)
    try:
        headers_row = next(rows_iter)
    except StopIteration:
        workbook.close()
        uploaded_file.seek(0)
        return [], []

    headers = [str(value) if value is not None else "" for value in headers_row]
    rows: list[dict[str, str]] = []
    for index, values in enumerate(rows_iter):
        row_map: dict[str, str] = {}
        for col_index, header in enumerate(headers):
            if not header:
                continue
            cell_value = None
            if values is not None and col_index < len(values):
                cell_value = values[col_index]
            row_map[header] = "" if cell_value is None else str(cell_value)
        rows.append(row_map)
        if index + 1 >= sample_rows:
            break

    workbook.close()
    uploaded_file.seek(0)
    return headers, rows


def _render_reembed_controls(group_id: str, sources: list[SourceFile]) -> None:
    st.subheader("Re-embed sources")
    options = [(source.id, source.version_label) for source in sources]
    selected = st.multiselect(
        "Select sources to re-embed",
        options=options,
        format_func=lambda item: f"{item[1]} ({item[0]})",
        key="reembed_select",
    )
    selected_ids = [item[0] for item in selected]
    if st.button("Queue re-embed", type="primary", disabled=not selected_ids, width="stretch"):
        job = default_queue.enqueue(group_id, selected_ids, triggered_by=None)
        record_audit(
            "ui.reembed",
            "queued",
            user=None,
            details={"group": group_id, "job": job.id, "sources": len(selected_ids)},
        )
        st.success(f"Queued re-embed job {job.id}")
    active_jobs = [default_queue.get_status(group_id, job_id) for job_id in [job.id for job in default_queue._completed.get(group_id, {}).values()]]  # noqa: SLF001
    if active_jobs:
        st.write("Recent jobs")
        st.dataframe(
            [
                {
                    "Job": job.id,
                    "Status": job.status.value,
                    "Completed": job.completed_at,
                    "Duration (ms)": job.run_duration_ms,
                }
                for job in active_jobs
                if job
            ],
            hide_index=True,
            width="stretch",
        )


def _render_preferences(group_id: str, sources: list[SourceFile]) -> None:
    st.subheader("Group preferences")
    available_columns: list[str] = []
    for source in sources:
        available_columns.extend(list(source.extracted_columns))
    available_columns = sorted({col for col in available_columns if col})
    prefs = st.session_state.setdefault("ingest_preferences", {})
    defaults = prefs.get(group_id, [])
    selection = st.multiselect(
        "Contextual columns",
        options=available_columns,
        default=defaults,
        help="Selections saved per group",
    )
    if st.button("Save preferences", type="primary"):
        prefs[group_id] = selection
        st.session_state["ingest_preferences"] = prefs
        record_audit("ui.preferences", "saved", user=None, details={"group": group_id, "count": len(selection)})
        st.success("Preferences saved for this group.")


def _collect_sheet_schemas(
    uploaded_file: UploadedFile, delimiter: str | None = None
) -> list[_UploadSheet]:
    """Read the uploaded file and return sheet-like metadata for catalog building."""
    uploaded_file.seek(0)
    filename = uploaded_file.name.lower()

    if filename.endswith(".csv"):
        content = uploaded_file.read().decode("utf-8")
        uploaded_file.seek(0)
        reader = csv.reader(io.StringIO(content), delimiter=delimiter or ",")
        try:
            headers = next(reader)
        except StopIteration:
            return []
        column_schema: list[dict[str, object]] = []
        for index, raw in enumerate(headers):
            name = str(raw or "").strip()
            if not name:
                name = f"Column {index + 1}"
                availability = "missing"
            else:
                availability = "available"
            column_schema.append({"name": name, "availability": availability})
        return [
            _UploadSheet(
                id="__csv__",
                display_label="CSV",
                status=SheetStatus.ACTIVE,
                column_schema=column_schema,
                last_refreshed_at=datetime.now(UTC),
            )
        ]

    if load_workbook is None:
        uploaded_file.seek(0)
        raise RuntimeError("Excel preview requires the 'openpyxl' dependency.")

    workbook = load_workbook(uploaded_file, read_only=True)
    sheets: list[_UploadSheet] = []
    try:
        for index, sheet_name in enumerate(workbook.sheetnames):
            worksheet = workbook[sheet_name]
            iterator = worksheet.iter_rows(values_only=True)
            try:
                headers_row = next(iterator)
            except StopIteration:
                headers_row = []

            column_schema: list[dict[str, object]] = []
            for col_index, value in enumerate(headers_row or []):
                header = str(value or "").strip()
                if header:
                    availability = "available"
                else:
                    header = f"Column {col_index + 1}"
                    availability = "missing"
                column_schema.append({"name": header, "availability": availability})

            sheets.append(
                _UploadSheet(
                    id=f"sheet-{index + 1}",
                    display_label=sheet_name,
                    status=SheetStatus.ACTIVE,
                    column_schema=column_schema,
                    last_refreshed_at=datetime.now(UTC),
                )
            )
    finally:
        workbook.close()
        uploaded_file.seek(0)

    return sheets


def _build_column_catalog(
    uploaded_file: UploadedFile, delimiter: str | None = None
) -> list[dict[str, object]]:
    sheets = _collect_sheet_schemas(uploaded_file, delimiter)
    if not sheets:
        return []
    catalog = aggregate_column_catalog(sheets, include_unavailable=True)
    return build_column_picker_options(catalog)


def _render_sheet_catalog(repo: MetadataRepository) -> None:
    bundles = repo.list_source_bundles()
    if not bundles:
        st.info("No source bundles ingested yet. Upload a workbook or CSV to begin.")
        return

    st.subheader("Sheet Source Catalog")
    catalog_rows: list[dict[str, object]] = []
    for bundle in bundles:
        sheets = repo.list_sheet_sources(bundle_id=bundle.id)
        if not sheets:
            continue
        embedding_counts = repo.get_sheet_embedding_counts(sheet_ids=[sheet.id for sheet in sheets])
        for sheet in sheets:
            vector_count = embedding_counts.get(sheet.id, 0)
            status_value = sheet.status.value
            status_label = status_value.replace("_", " ").title()
            visibility_label = (
                "Hidden opt-in" if sheet.visibility_state.value == "hidden_opt_in" else "Visible"
            )
            needs_attention = sheet.status != SheetStatus.ACTIVE or vector_count == 0
            catalog_rows.append(
                {
                    "Bundle": bundle.display_name,
                    "Sheet": sheet.display_label,
                    "Visibility": visibility_label,
                    "Status": status_label,
                    "Rows": sheet.row_count,
                    "Embeddings": vector_count,
                    "Last refreshed": (
                        sheet.last_refreshed_at.isoformat() if sheet.last_refreshed_at else "N/A"
                    ),
                    "_status_value": status_value,
                    "_needs_attention": needs_attention,
                    "_text_blob": f"{bundle.display_name} {sheet.display_label} {visibility_label}".lower(),
                }
            )

    if not catalog_rows:
        st.info("Source bundles exist but no sheet sources have been registered yet.")
        return

    attention_total = sum(1 for row in catalog_rows if row["_needs_attention"])
    col_a, col_b, col_c = st.columns(3)
    col_a.metric("Bundles", len(bundles))
    col_b.metric("Sheet sources", len(catalog_rows))
    col_c.metric("Needs attention", attention_total)

    st.caption("Use the interactive controls below to search, filter, and spotlight sheets quickly.")

    search_query = st.text_input(
        "Search bundles or sheets",
        placeholder="Type a bundle or sheet name…",
        key="sheet_catalog_search",
    ).strip()
    status_options = sorted({row["Status"] for row in catalog_rows})
    selected_statuses = st.multiselect(
        "Status filter",
        options=status_options,
        default=status_options,
        key="sheet_catalog_status_filter",
    )
    attention_only = st.checkbox(
        "Show only sheets needing attention (inactive or missing embeddings)",
        key="sheet_catalog_attention_only",
        value=False,
    )

    lowered_query = search_query.lower()
    filtered_rows: list[dict[str, object]] = []
    for row in catalog_rows:
        if selected_statuses and row["Status"] not in selected_statuses:
            continue
        if lowered_query and lowered_query not in row["_text_blob"]:
            continue
        if attention_only and not row["_needs_attention"]:
            continue
        filtered_rows.append(row)

    filtered_rows.sort(key=lambda row: (row["Bundle"].lower(), row["Sheet"].lower()))
    display_rows = [
        {
            "Bundle": row["Bundle"],
            "Sheet": row["Sheet"],
            "Visibility": row["Visibility"],
            "Status": row["Status"],
            "Rows": row["Rows"],
            "Embeddings": row["Embeddings"],
            "Last refreshed": row["Last refreshed"],
        }
        for row in filtered_rows
    ]

    if not display_rows:
        st.warning("No sheet sources match the current filters.")
        return

    st.caption(f"Showing {len(display_rows)} of {len(catalog_rows)} sheet sources.")
    st.data_editor(
        display_rows,
        hide_index=True,
        width="stretch",
        disabled=True,
        column_config={
            "Embeddings": st.column_config.NumberColumn(format="%d", help="Embeddings currently stored"),
            "Rows": st.column_config.NumberColumn(format="%d"),
        },
    )


def run_ingestion(
    uploaded_file: UploadedFile,
    selected_columns: Sequence[str],
    delimiter: str | None,
    hidden_sheet_overrides: Sequence[str],
) -> None:
    session_factory = _get_session_factory()

    with session_scope(session_factory) as session:
        repo = MetadataRepository(session)
        embedding_service = EmbeddingService(metadata_repository=repo)
        ingestion_service = IngestionService(
            metadata_repository=repo,
            embedding_service=embedding_service,
            data_root=_data_root(),
        )

        with tempfile.NamedTemporaryFile(
            delete=False, suffix=Path(uploaded_file.name).suffix
        ) as tmp:
            tmp.write(uploaded_file.read())
            temp_path = Path(tmp.name)
        uploaded_file.seek(0)

        try:
            with log_timing(LOGGER, "streamlit.ingestion.submit", display_name=uploaded_file.name):
                bundle_result = ingestion_service.ingest_bundle(
                    source_path=temp_path,
                    display_name=uploaded_file.name,
                    options=BundleIngestionOptions(
                        selected_columns=selected_columns,
                        delimiter=delimiter or None,
                        hidden_sheet_policy=HiddenSheetPolicy(
                            default_action="exclude",
                            overrides=list(hidden_sheet_overrides),
                        ),
                    ),
                )
        finally:
            temp_path.unlink(missing_ok=True)

        log_event(LOGGER, "streamlit.ingestion.complete", bundle_id=bundle_result.bundle.id)
        return bundle_result


def main() -> None:
    st.title("Ingest Sheet Sources")
    st.caption(
        "Upload CSV or Excel files, expose individual sheets to the catalog, "
        "and opt in hidden tabs with audit logging."
    )
    _render_persistence_status()
    group_id, sources = _render_source_manager()
    _render_reembed_controls(group_id, sources)
    _render_preferences(group_id, sources)
    st.divider()

    state = ensure_session_defaults()
    reset_flag = "ingest_reset_pending"
    if st.button("Reset saved column selections", type="secondary", key="ingest_reset_button"):
        request_reset(state, reason="clear ingest selections")
        st.session_state[reset_flag] = True
    if st.session_state.get(reset_flag):
        st.warning(
            "Reset clears saved selections for this session before the next upload.",
            icon="⚠️",
        )
        if st.button("Confirm reset", type="primary", key="ingest_reset_confirm"):
            confirm_reset(state)
            st.session_state.pop("local_preference_payload", None)
            st.session_state.pop(reset_flag, None)
            st.success("Selections cleared. Upload a file to start fresh.")

    uploaded_file = st.file_uploader("Select CSV or Excel file", type=["csv", "xlsx", "xls"])
    selected_columns: list[str] = list(state.get("selected_columns", []))
    delimiter: str | None = None
    preview_rows: list[dict[str, str]] = []
    picker_options: list[dict[str, object]] = []
    hidden_sheet_overrides: list[str] = []
    visible_sheet_names: list[str] = []
    hidden_sheet_names: list[str] = []

    if uploaded_file:
        try:
            filename = uploaded_file.name.lower()
            if filename.endswith(".csv"):
                delimiter = st.text_input("Delimiter", value=",")

            _headers, preview_rows = _preview_rows(uploaded_file)
            if preview_rows:
                st.table(preview_rows)

            picker_options = _build_column_catalog(uploaded_file, delimiter)
            available_options = [
                option for option in picker_options if option["availability"] == "available"
            ]
            unavailable_options = [
                option for option in picker_options if option["availability"] != "available"
            ]

            display_labels = {
                option["column_name"]: (
                    f"{option['display_label']} ({', '.join(option['sheet_chips'])})"
                    if option.get("sheet_chips")
                    else option["display_label"]
                )
                for option in available_options
            }
            available_names = list(display_labels.keys())
            default_selection = [name for name in selected_columns if name in display_labels]

            selected_columns = st.multiselect(
                "Select text columns to embed",
                options=available_names,
                default=default_selection,
                format_func=lambda name: display_labels.get(name, name),
                help="Columns are deduped across sheets. Use arrow keys and Enter to pick quickly.",
            )
            persist_column_selection(
                st.session_state,
                dataset_id=uploaded_file.name,
                selected_columns=selected_columns,
                active_tab="ingest",
            )

            if unavailable_options:
                st.warning(
                    "Unavailable or missing headers are skipped from selection but listed below."
                )
                for option in unavailable_options:
                    chips = ", ".join(option.get("sheet_chips", []))
                    st.caption(f"- {option['display_label']} ({chips}) — {option['availability']}")

            if filename.endswith((".xls", ".xlsx")):
                if load_workbook is None:
                    st.error("Excel ingestion requires installing the 'openpyxl' package.")
                else:
                    workbook = load_workbook(uploaded_file, read_only=True)
                    visible_sheet_names = []
                    hidden_sheet_names = []
                    for name in workbook.sheetnames:
                        worksheet = workbook[name]
                        hidden = getattr(worksheet, "sheet_state", "visible") != "visible"
                        if hidden:
                            hidden_sheet_names.append(name)
                        else:
                            visible_sheet_names.append(name)
                    workbook.close()
                    uploaded_file.seek(0)

                    if visible_sheet_names:
                        st.info("Visible sheets detected: " + ", ".join(visible_sheet_names))
                    if hidden_sheet_names:
                        st.warning(
                            "Hidden sheets stay excluded unless explicitly enabled. "
                            "Select hidden sheets to include; this will be recorded in audit logs."
                        )
                        hidden_sheet_overrides = st.multiselect(
                            "Hidden sheets to include",
                            options=hidden_sheet_names,
                        )
            else:
                visible_sheet_names = ["__csv__"]
        except Exception as error:
            st.error(f"Failed to read file: {error}")
            selected_columns = []
            hidden_sheet_overrides = []

    disable_ingest = True
    if uploaded_file and selected_columns:
        if uploaded_file.name.lower().endswith((".xls", ".xlsx")) and load_workbook is None:
            disable_ingest = True
        else:
            disable_ingest = False

    if st.button("Start Ingestion", disabled=disable_ingest):
        try:
            bundle_result = run_ingestion(
                uploaded_file,
                selected_columns,
                delimiter,
                hidden_sheet_overrides,
            )
            st.success(
                f"Ingestion complete: bundle `{bundle_result.bundle.display_name}` "
                f"ready with {bundle_result.bundle.sheet_count} sheet sources."
            )

            sheet_rows = [
                {
                    "Sheet": sheet_result.sheet.sheet_name,
                    "Visibility": sheet_result.sheet.visibility_state.value,
                    "Rows": sheet_result.sheet.row_count,
                    "Embeddings": sheet_result.embedding_summary.vector_count,
                }
                for sheet_result in bundle_result.sheets
            ]
            if sheet_rows:
                st.table(sheet_rows)

            if hidden_sheet_overrides:
                st.info("Hidden sheet opt-ins recorded: " + ", ".join(hidden_sheet_overrides))
        except Exception as error:
            st.error(f"Ingestion failed: {error}")

    with session_scope(_get_session_factory()) as session:
        repo = MetadataRepository(session)
        _render_sheet_catalog(repo)


if __name__ == "__main__":
    main()
