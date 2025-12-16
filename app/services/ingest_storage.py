from __future__ import annotations

import csv
import io
import json
import uuid
from collections.abc import Iterable, Sequence
from datetime import UTC, datetime
from pathlib import Path
from typing import BinaryIO

import pandas as pd

try:
    from openpyxl import load_workbook
except ImportError:  # pragma: no cover - optional
    load_workbook = None

from app.services.ingest_models import SourceFile, SourceStatus
from app.utils.config import IngestConfig, load_ingest_config
from app.utils.logging import get_logger

LOGGER = get_logger(__name__)


class IngestStorage:
    """Manage ingest source files with validation, versioning, and column extraction."""

    def __init__(self, config: IngestConfig | None = None) -> None:
        self.config = config or load_ingest_config()
        self.storage_root = self.config.storage_root
        self.storage_root.mkdir(parents=True, exist_ok=True)

    # ---- metadata helpers ----
    def _index_path(self, group_id: str) -> Path:
        return self.storage_root / group_id / "_index.json"

    def _load_index(self, group_id: str) -> list[dict]:
        index_path = self._index_path(group_id)
        if not index_path.exists():
            return []
        try:
            return json.loads(index_path.read_text())
        except Exception:  # pragma: no cover - corrupt index
            LOGGER.warning("Corrupt ingest index for %s, rebuilding", group_id)
            return []

    def _save_index(self, group_id: str, entries: Sequence[dict]) -> None:
        index_path = self._index_path(group_id)
        index_path.parent.mkdir(parents=True, exist_ok=True)
        index_path.write_text(json.dumps(list(entries), indent=2, default=str))

    # ---- validation / naming ----
    def _allowed_extension(self, name: str) -> bool:
        lower = name.lower()
        for ext in self.config.allowed_types:
            if lower.endswith(f".{ext}"):
                return True
        return False

    def _version_label(self, group_dir: Path, filename: str) -> str:
        base_path = group_dir / filename
        if not base_path.exists():
            return filename
        stem = base_path.stem
        suffix = base_path.suffix
        counter = 2
        while True:
            candidate = group_dir / f"{stem} ({counter}){suffix}"
            if not candidate.exists():
                return candidate.name
            counter += 1

    def _validate(self, name: str, content: bytes) -> tuple[str, int]:
        size = len(content)
        if size > self.config.max_bytes:
            raise ValueError(f"File too large: {size} > {self.config.max_bytes}")
        if not self._allowed_extension(name):
            raise ValueError(f"Unsupported file type for {name}")
        return name, size

    # ---- public API ----
    def list_sources(self, group_id: str) -> list[SourceFile]:
        sources: list[SourceFile] = []
        for entry in self._load_index(group_id):
            try:
                sources.append(self._deserialize_source(entry))
            except Exception:
                continue
        return sources

    def save_upload(
        self,
        group_id: str,
        uploaded_file: BinaryIO,
        *,
        filename: str,
        mime_type: str,
        added_by: str | None = None,
    ) -> SourceFile:
        group_dir = self.storage_root / group_id
        group_dir.mkdir(parents=True, exist_ok=True)

        content = uploaded_file.read()
        uploaded_file.seek(0)
        safe_name, size = self._validate(filename, content)
        version_label = self._version_label(group_dir, safe_name)
        target_path = group_dir / version_label
        target_path.write_bytes(content)

        extracted_columns = tuple(self.extract_columns(target_path))
        source = SourceFile(
            id=str(uuid.uuid4()),
            document_group_id=group_id,
            filename=safe_name,
            version_label=version_label,
            size_bytes=size,
            mime_type=mime_type,
            storage_path=str(target_path),
            added_by=added_by,
            status=SourceStatus.READY,
            last_updated_at=datetime.now(UTC),
            extracted_columns=extracted_columns,
        )

        index = self._load_index(group_id)
        index.append(self._serialize_source(source))
        self._save_index(group_id, index)
        return source

    def delete_source(self, group_id: str, source_id: str) -> bool:
        index = self._load_index(group_id)
        remaining: list[dict] = []
        deleted = False
        for entry in index:
            if entry.get("id") != source_id:
                remaining.append(entry)
                continue
            deleted = True
            path = entry.get("storage_path")
            if path:
                Path(path).unlink(missing_ok=True)
        if deleted:
            self._save_index(group_id, remaining)
        return deleted

    def extract_columns(self, file_path: Path) -> list[str]:
        suffix = file_path.suffix.lower()
        if suffix == ".csv":
            return self._extract_csv(file_path)
        if suffix in {".xlsx", ".xls"}:
            return self._extract_excel(file_path)
        if suffix == ".parquet":
            return self._extract_parquet(file_path)
        return []

    # ---- extraction helpers ----
    def _extract_csv(self, file_path: Path) -> list[str]:
        with file_path.open("r", encoding="utf-8", newline="") as handle:
            reader = csv.reader(handle)
            try:
                headers = next(reader)
            except StopIteration:
                return []
        return _unique_nonempty(headers)

    def _extract_excel(self, file_path: Path) -> list[str]:
        if load_workbook is None:
            return []
        workbook = load_workbook(file_path, read_only=True)
        headers: list[str] = []
        try:
            for name in workbook.sheetnames:
                sheet = workbook[name]
                iterator = sheet.iter_rows(values_only=True)
                try:
                    first_row = next(iterator)
                except StopIteration:
                    continue
                headers.extend([value for value in first_row if value is not None])
        finally:
            workbook.close()
        return _unique_nonempty(headers)

    def _extract_parquet(self, file_path: Path) -> list[str]:
        try:
            df = pd.read_parquet(file_path, engine="pyarrow")
        except Exception:  # pragma: no cover - fallback path
            return []
        return _unique_nonempty(df.columns)

    # ---- serialization helpers ----
    @staticmethod
    def _serialize_source(source: SourceFile) -> dict:
        return {
            "id": source.id,
            "document_group_id": source.document_group_id,
            "filename": source.filename,
            "version_label": source.version_label,
            "size_bytes": source.size_bytes,
            "mime_type": source.mime_type,
            "storage_path": source.storage_path,
            "added_by": source.added_by,
            "added_at": source.added_at.isoformat(),
            "status": source.status.value,
            "last_updated_at": source.last_updated_at.isoformat() if source.last_updated_at else None,
            "validation_error": source.validation_error,
            "audit_log_ref": source.audit_log_ref,
            "extracted_columns": list(source.extracted_columns),
        }

    @staticmethod
    def _deserialize_source(payload: dict) -> SourceFile:
        return SourceFile(
            id=payload["id"],
            document_group_id=payload["document_group_id"],
            filename=payload["filename"],
            version_label=payload["version_label"],
            size_bytes=int(payload["size_bytes"]),
            mime_type=payload.get("mime_type") or "",
            storage_path=payload["storage_path"],
            added_by=payload.get("added_by"),
            added_at=datetime.fromisoformat(payload["added_at"]),
            status=SourceStatus(payload.get("status", SourceStatus.READY)),
            last_updated_at=datetime.fromisoformat(payload["last_updated_at"])
            if payload.get("last_updated_at")
            else None,
            validation_error=payload.get("validation_error"),
            audit_log_ref=payload.get("audit_log_ref"),
            extracted_columns=tuple(payload.get("extracted_columns") or ()),
        )

    # ---- preference helpers ----
    def _pref_path(self, group_id: str) -> Path:
        root = self.storage_root / group_id
        root.mkdir(parents=True, exist_ok=True)
        return root / "_preferences.json"

    def save_preferences(self, group_id: str, selected_columns: Sequence[str], contextual_fields: Sequence[str] | None) -> dict:
        payload = {
            "selected_columns": list(selected_columns),
            "contextual_fields": list(contextual_fields or []),
        }
        self._pref_path(group_id).write_text(json.dumps(payload, indent=2))
        return payload

    def load_preferences(self, group_id: str) -> dict:
        path = self._pref_path(group_id)
        if not path.exists():
            return {"selected_columns": [], "contextual_fields": []}
        try:
            return json.loads(path.read_text())
        except Exception:  # pragma: no cover - corrupt preferences
            LOGGER.warning("Failed to read preferences for %s", group_id)
            return {"selected_columns": [], "contextual_fields": []}


def _unique_nonempty(headers: Iterable[str]) -> list[str]:
    seen: set[str] = set()
    result: list[str] = []
    for raw in headers:
        name = str(raw or "").strip()
        if not name:
            continue
        if name in seen:
            continue
        seen.add(name)
        result.append(name)
    return result


default_storage = IngestStorage()
