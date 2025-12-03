from __future__ import annotations

import csv
import hashlib
import json
import shutil
import time
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Literal, Optional, Sequence

try:  # Optional dependency for Excel ingestion
    from openpyxl import load_workbook
except ImportError:  # pragma: no cover - optional path
    load_workbook = None

from app.db.metadata import MetadataRepository
from app.db.schema import (
    AuditStatus,
    BundleAudit,
    DataFile,
    FileType,
    IngestionStatus,
    MetricType,
    QueryRecord,
    SheetSource,
    SheetStatus,
    SheetVisibilityState,
    SourceBundle,
)
from app.services.embeddings import EmbeddingJob, EmbeddingService, EmbeddingSummary
from app.services.refresh_utils import (
    DiscoveredSheetSnapshot,
    ExistingSheetSnapshot,
    match_sheets,
)
from app.utils.logging import get_logger, log_timing

LOGGER = get_logger(__name__)


@dataclass(frozen=True)
class DiscoveredSheet:
    name: str
    position: int
    hidden: bool


@dataclass(frozen=True)
class HiddenSheetPolicy:
    default_action: Literal["exclude", "include_all"] = "exclude"
    overrides: Sequence[str] = ()

    def allows(self, sheet: DiscoveredSheet) -> bool:
        if not sheet.hidden:
            return True
        if self.default_action == "include_all":
            return True
        override_names = {override.lower() for override in self.overrides}
        return sheet.name.lower() in override_names


@dataclass(frozen=True)
class BundleIngestionOptions:
    selected_columns: Sequence[str]
    hidden_sheet_policy: HiddenSheetPolicy
    delimiter: Optional[str] = None
    encoding: str = "utf-8"
    allow_duplicate_import: bool = False


@dataclass
class SheetIngestionResult:
    sheet: SheetSource
    data_file: DataFile
    embedding_summary: EmbeddingSummary
    processed_rows: int
    skipped_rows: int


@dataclass
class BundleIngestionResult:
    bundle: SourceBundle
    sheets: list[SheetIngestionResult]
    hidden_opt_ins: list[DiscoveredSheet]
    audit: BundleAudit


@dataclass
class BundleRefreshResult:
    bundle: SourceBundle
    created: list[SheetIngestionResult]
    updated: list[SheetIngestionResult]
    deactivated: list[SheetSource]
    audit: BundleAudit


@dataclass(frozen=True)
class ColumnCatalogEntry:
    column_name: str
    display_label: str
    availability: Literal["available", "missing", "unavailable"]
    sheet_ids: tuple[str, ...]
    sheet_labels: tuple[str, ...]
    data_type: str | None
    last_seen_at: datetime | None
    normalized_key: str


def _normalize_column_key(value: str) -> str:
    return " ".join(value.lower().split())


def _merge_availability(
    current: Literal["available", "missing", "unavailable"],
    candidate: Literal["available", "missing", "unavailable"],
) -> Literal["available", "missing", "unavailable"]:
    priority = {"available": 2, "missing": 1, "unavailable": 0}
    return current if priority[current] >= priority[candidate] else candidate


def aggregate_column_catalog(
    sheets: Sequence[SheetSource],
    *,
    include_unavailable: bool = False,
) -> list[ColumnCatalogEntry]:
    """Build a deduplicated column catalog with availability and sheet provenance."""
    catalog: dict[str, dict[str, object]] = {}

    for sheet in sheets:
        if sheet.status != SheetStatus.ACTIVE:
            continue
        schema = getattr(sheet, "column_schema", None) or []
        latest_seen = sheet.last_refreshed_at
        seen_keys: set[str] = set()

        for column in schema:
            raw_name = str(column.get("name", "")).strip()
            if not raw_name:
                continue
            normalized = _normalize_column_key(raw_name)
            if not normalized or normalized in seen_keys:
                continue
            seen_keys.add(normalized)

            availability = str(column.get("availability", "available")).lower()
            if availability not in {"available", "missing", "unavailable"}:
                availability = "available"
            data_type = column.get("inferredType") or column.get("data_type")
            display_label = str(column.get("display_label", raw_name)).strip() or raw_name

            entry = catalog.get(normalized)
            if entry is None:
                entry = {
                    "column_name": raw_name,
                    "display_label": display_label,
                    "availability": availability,
                    "data_type": str(data_type) if data_type else None,
                    "sheet_ids": set(),
                    "sheet_labels": [],
                    "last_seen_at": latest_seen,
                }
                catalog[normalized] = entry
            else:
                entry["availability"] = _merge_availability(
                    entry["availability"], availability  # type: ignore[arg-type]
                )
                if entry["data_type"] is None and data_type:
                    entry["data_type"] = str(data_type)
                if entry["last_seen_at"] is None or (
                    latest_seen and latest_seen > entry["last_seen_at"]
                ):
                    entry["last_seen_at"] = latest_seen

            if sheet.id not in entry["sheet_ids"]:
                entry["sheet_ids"].add(sheet.id)  # type: ignore[attr-defined]
            if sheet.display_label not in entry["sheet_labels"]:
                entry["sheet_labels"].append(sheet.display_label)  # type: ignore[attr-defined]

    entries: list[ColumnCatalogEntry] = []
    for normalized, entry in catalog.items():
        availability = entry["availability"]  # type: ignore[assignment]
        if not include_unavailable and availability != "available":
            continue
        entries.append(
            ColumnCatalogEntry(
                column_name=entry["column_name"],  # type: ignore[arg-type]
                display_label=entry["display_label"],  # type: ignore[arg-type]
                availability=availability,  # type: ignore[arg-type]
                sheet_ids=tuple(sorted(entry["sheet_ids"])),  # type: ignore[arg-type]
                sheet_labels=tuple(entry["sheet_labels"]),  # type: ignore[arg-type]
                data_type=entry["data_type"],  # type: ignore[arg-type]
                last_seen_at=entry["last_seen_at"],  # type: ignore[arg-type]
                normalized_key=normalized,
            )
        )
    entries.sort(key=lambda item: item.display_label.lower())
    return entries


def build_column_picker_options(
    catalog: Sequence[ColumnCatalogEntry],
) -> list[dict[str, object]]:
    """Prepare a UI-friendly representation of catalog entries with sheet chips."""
    options: list[dict[str, object]] = []
    for entry in catalog:
        sheet_chips = sorted(set(entry.sheet_labels)) if entry.sheet_labels else sorted(entry.sheet_ids)
        options.append(
            {
                "column_name": entry.column_name,
                "display_label": entry.display_label,
                "availability": entry.availability,
                "sheet_chips": sheet_chips,
                "data_type": entry.data_type,
                "last_seen_at": entry.last_seen_at,
            }
        )
    return options


@dataclass
class IngestionOptions:
    selected_columns: Sequence[str]
    delimiter: Optional[str] = None
    sheet_name: Optional[str] = None
    encoding: str = "utf-8"
    allow_duplicate_import: bool = False


@dataclass
class IngestionResult:
    data_file: DataFile
    embedding_summary: EmbeddingSummary
    processed_rows: int
    skipped_rows: int


def discover_workbook_sheets(path: Path) -> list[DiscoveredSheet]:
    if load_workbook is None:
        raise RuntimeError("Excel ingestion requires the 'openpyxl' dependency.")
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        discovered: list[DiscoveredSheet] = []
        for position, name in enumerate(workbook.sheetnames):
            worksheet = workbook[name]
            hidden = getattr(worksheet, "sheet_state", "visible") != "visible"
            discovered.append(DiscoveredSheet(name=name, position=position, hidden=hidden))
        return discovered
    finally:
        workbook.close()


def apply_hidden_sheet_policy(
    sheets: Sequence[DiscoveredSheet],
    policy: HiddenSheetPolicy,
) -> tuple[list[DiscoveredSheet], list[DiscoveredSheet], list[DiscoveredSheet]]:
    included: list[DiscoveredSheet] = []
    hidden_opt_ins: list[DiscoveredSheet] = []
    excluded: list[DiscoveredSheet] = []
    override_names = {override.lower() for override in policy.overrides}

    for sheet in sheets:
        if not sheet.hidden:
            included.append(sheet)
            continue
        if policy.default_action == "include_all" or sheet.name.lower() in override_names:
            included.append(sheet)
            hidden_opt_ins.append(sheet)
        else:
            excluded.append(sheet)
    return included, hidden_opt_ins, excluded


def build_sheet_summary(
    created: Sequence[DiscoveredSheet],
    hidden_opt_ins: Sequence[DiscoveredSheet],
    inactive: int,
) -> dict[str, int]:
    return {
        "created": len(created),
        "hidden_opt_ins": len(hidden_opt_ins),
        "inactive": inactive,
    }


class IngestionService:
    def __init__(
        self,
        *,
        metadata_repository: MetadataRepository,
        embedding_service: EmbeddingService,
        data_root: Path | str,
    ) -> None:
        self.metadata_repository = metadata_repository
        self.embedding_service = embedding_service
        self.data_root = Path(data_root)
        self.data_root.mkdir(parents=True, exist_ok=True)
        (self.data_root / "raw").mkdir(parents=True, exist_ok=True)
        (self.data_root / "bundles").mkdir(parents=True, exist_ok=True)

    def ingest_file(
        self,
        *,
        source_path: Path,
        display_name: str,
        options: IngestionOptions,
    ) -> IngestionResult:
        if not options.selected_columns:
            raise ValueError("At least one column must be selected for ingestion.")
        if not source_path.exists():
            raise FileNotFoundError(f"Source file does not exist: {source_path}")

        start_time = datetime.now(timezone.utc)
        timer = time.perf_counter()

        file_hash = self._hash_file(source_path)
        existing = self.metadata_repository.get_data_file_by_hash(file_hash)
        if existing and not options.allow_duplicate_import:
            raise ValueError("Dataset with identical content already ingested.")

        file_type = self._infer_file_type(source_path)
        data_file = self.metadata_repository.create_data_file(
            display_name=display_name,
            original_path=str(source_path.resolve()),
            file_hash=file_hash,
            file_type=file_type,
            delimiter=options.delimiter,
            sheet_name=options.sheet_name,
            selected_columns=options.selected_columns,
            status=IngestionStatus.PROCESSING,
        )
        self.metadata_repository.session.flush()  # type: ignore[attr-defined]

        copied_path = self._copy_raw_file(source_path, data_file)

        try:
            with log_timing(LOGGER, "ingestion.load_records", dataset_id=data_file.id):
                columns, rows = self._load_rows(copied_path, file_type, options)
            (
                records,
                skipped_rows,
                available_columns,
                missing_columns,
            ) = self._materialize_records(
                rows=rows,
                columns=columns,
                data_file=data_file,
                selected_columns=options.selected_columns,
            )
            if not available_columns:
                raise ValueError(
                    "Selected columns not found in source; ingestion requires at least one matching column."
                )

            data_file.selected_columns = list(available_columns)
            if missing_columns:
                LOGGER.warning(
                    "Some selected columns were not found for dataset %s and will be skipped: %s",
                    data_file.id,
                    missing_columns,
                )

            processed_rows = len(records)
            if processed_rows == 0:
                raise ValueError("No valid textual rows found for the selected columns.")

            embedding_summary = self.embedding_service.run_embedding(
                EmbeddingJob(
                    data_file=data_file,
                    records=records,
                    metadata_repository=self.metadata_repository,
                )
            )

            self.metadata_repository.update_data_file_status(
                data_file,
                status=IngestionStatus.READY,
                row_count=processed_rows,
                error_summary=None,
                processed_at=datetime.now(timezone.utc),
            )

            elapsed_ms = (time.perf_counter() - timer) * 1000.0
            records_per_second = processed_rows / (elapsed_ms / 1000.0) if elapsed_ms else None

            self.metadata_repository.create_audit(
                data_file_id=data_file.id,
                status=AuditStatus.SUCCEEDED,
                processed_rows=processed_rows,
                skipped_rows=skipped_rows,
                started_at=start_time,
                completed_at=datetime.now(timezone.utc),
            )

            self.metadata_repository.record_performance_metric(
                metric_type=MetricType.INGESTION,
                data_file_id=data_file.id,
                cluster_id=None,
                benchmark_run_id=None,
                p50_ms=elapsed_ms,
                p95_ms=elapsed_ms,
                records_per_second=records_per_second,
            )

            self.metadata_repository.session.commit()  # type: ignore[attr-defined]

            return IngestionResult(
                data_file=data_file,
                embedding_summary=embedding_summary,
                processed_rows=processed_rows,
                skipped_rows=skipped_rows,
            )
        except Exception as error:
            LOGGER.exception("Ingestion failed for dataset %s: %s", data_file.id, error)
            self.metadata_repository.update_data_file_status(
                data_file,
                status=IngestionStatus.FAILED,
                error_summary=str(error),
            )
            self.metadata_repository.create_audit(
                data_file_id=data_file.id,
                status=AuditStatus.FAILED,
                processed_rows=0,
                skipped_rows=0,
                started_at=start_time,
                completed_at=datetime.now(timezone.utc),
            )
            self.metadata_repository.session.commit()  # type: ignore[attr-defined]
            raise

    def ingest_bundle(
        self,
        *,
        source_path: Path,
        display_name: str,
        options: BundleIngestionOptions,
    ) -> BundleIngestionResult:
        if not options.selected_columns:
            raise ValueError("At least one column must be selected for ingestion.")
        if not source_path.exists():
            raise FileNotFoundError(f"Source file does not exist: {source_path}")

        ingestion_started = datetime.now(timezone.utc)
        file_hash = self._hash_file(source_path)
        existing_bundle = self.metadata_repository.get_source_bundle_by_hash(file_hash)
        if existing_bundle and not options.allow_duplicate_import:
            raise ValueError("Bundle with identical content already ingested.")

        file_type = self._infer_file_type(source_path)
        bundle = self.metadata_repository.create_source_bundle(
            display_name=display_name,
            original_path=str(source_path.resolve()),
            file_hash=file_hash,
            file_type=file_type,
            delimiter=options.delimiter if file_type == FileType.CSV else None,
            refresh_cadence=None,
            ingestion_status=IngestionStatus.PROCESSING,
        )
        self.metadata_repository.session.flush()  # type: ignore[attr-defined]

        stored_path = self._copy_bundle_file(source_path, bundle)
        self.metadata_repository.update_source_bundle(
            bundle,
            original_path=str(stored_path),
        )

        if file_type == FileType.EXCEL:
            discovered = discover_workbook_sheets(stored_path)
        else:
            discovered = [DiscoveredSheet(name="__csv__", position=0, hidden=False)]

        included, hidden_opt_ins, _excluded = apply_hidden_sheet_policy(
            discovered, options.hidden_sheet_policy
        )
        if not included:
            raise ValueError("No sheets available for ingestion after applying hidden sheet policy.")

        sheet_results: list[SheetIngestionResult] = []
        inactive_sheets = 0

        try:
            for sheet in included:
                sheet_timer = time.perf_counter()
                sheet_started = datetime.now(timezone.utc)
                sheet_name = "__csv__" if file_type == FileType.CSV else sheet.name
                visibility_state = (
                    SheetVisibilityState.HIDDEN_OPT_IN if sheet.hidden else SheetVisibilityState.VISIBLE
                )
                position_index = sheet.position
                display_label = (
                    display_name if file_type == FileType.CSV else f"{display_name}:{sheet_name}"
                )

                sheet_model = self.metadata_repository.create_sheet_source(
                    bundle=bundle,
                    sheet_name=sheet_name,
                    display_label=display_label,
                    visibility_state=visibility_state,
                    status=SheetStatus.ACTIVE,
                    row_count=0,
                    column_schema=[],
                    position_index=position_index,
                    checksum=None,
                )
                self.metadata_repository.session.flush()  # type: ignore[attr-defined]

                sheet_options = IngestionOptions(
                    selected_columns=options.selected_columns,
                    delimiter=options.delimiter,
                    sheet_name=None if file_type == FileType.CSV else sheet_name,
                    encoding=options.encoding,
                )
                columns, rows = self._load_rows(stored_path, file_type, sheet_options)
                row_count = len(rows)

                if row_count == 0:
                    self.metadata_repository.update_sheet_source(
                        sheet_model,
                        status=SheetStatus.INACTIVE,
                        row_count=0,
                        column_schema=[],
                        checksum=None,
                        last_refreshed_at=datetime.now(timezone.utc),
                        position_index=position_index,
                    )
                    inactive_sheets += 1
                    continue

                sheet_checksum = self._hash_sheet_content(columns=columns, rows=rows)
                sheet_hash = self._hash_sheet_identifier(bundle.file_hash, sheet_name)

                data_file = self.metadata_repository.create_data_file(
                    display_name=display_label,
                    original_path=str(stored_path),
                    file_hash=sheet_hash,
                    file_type=file_type,
                    delimiter=options.delimiter if file_type == FileType.CSV else None,
                    sheet_name=sheet_name if file_type == FileType.EXCEL else "__csv__",
                    selected_columns=options.selected_columns,
                    status=IngestionStatus.PROCESSING,
                )
                self.metadata_repository.session.flush()  # type: ignore[attr-defined]

                (
                    records,
                    skipped_rows,
                    available_columns,
                    missing_columns,
                ) = self._materialize_records(
                    rows=rows,
                    columns=columns,
                    data_file=data_file,
                    selected_columns=options.selected_columns,
                    sheet_id=sheet_model.id,
                )
                if not available_columns:
                    LOGGER.warning(
                        "Sheet '%s' missing all selected columns; marking inactive and continuing.",
                        sheet_name,
                    )
                    now = datetime.now(timezone.utc)
                    self.metadata_repository.update_data_file_status(
                        data_file,
                        status=IngestionStatus.FAILED,
                        row_count=row_count,
                        error_summary="Selected columns not present on sheet.",
                        processed_at=now,
                    )
                    self.metadata_repository.create_audit(
                        data_file_id=data_file.id,
                        status=AuditStatus.FAILED,
                        processed_rows=0,
                        skipped_rows=0,
                        started_at=sheet_started,
                        completed_at=now,
                    )
                    self.metadata_repository.update_sheet_source(
                        sheet_model,
                        status=SheetStatus.INACTIVE,
                        row_count=row_count,
                        column_schema=self._build_column_schema(columns=columns, rows=rows),
                        checksum=sheet_checksum,
                        last_refreshed_at=now,
                        position_index=position_index,
                    )
                    inactive_sheets += 1
                    continue

                data_file.selected_columns = list(available_columns)
                if missing_columns:
                    LOGGER.warning(
                        "Sheet '%s' missing selected columns; continuing with available subset: %s",
                        sheet_name,
                        missing_columns,
                    )

                processed_rows = len(records)
                if processed_rows == 0:
                    raise ValueError(f"No valid textual rows found for sheet '{sheet_name}'.")

                embedding_summary = self.embedding_service.run_embedding(
                    EmbeddingJob(
                        data_file=data_file,
                        sheet=sheet_model,
                        records=records,
                        metadata_repository=self.metadata_repository,
                    )
                )

                completed_at = datetime.now(timezone.utc)
                self.metadata_repository.update_data_file_status(
                    data_file,
                    status=IngestionStatus.READY,
                    row_count=row_count,
                    error_summary=None,
                    processed_at=completed_at,
                )

                elapsed_ms = (time.perf_counter() - sheet_timer) * 1000.0
                records_per_second = processed_rows / (elapsed_ms / 1000.0) if elapsed_ms else None

                self.metadata_repository.create_audit(
                    data_file_id=data_file.id,
                    status=AuditStatus.SUCCEEDED,
                    processed_rows=processed_rows,
                    skipped_rows=skipped_rows,
                    started_at=sheet_started,
                    completed_at=completed_at,
                )

                self.metadata_repository.record_performance_metric(
                    metric_type=MetricType.INGESTION,
                    data_file_id=data_file.id,
                    cluster_id=None,
                    benchmark_run_id=None,
                    p50_ms=elapsed_ms,
                    p95_ms=elapsed_ms,
                    records_per_second=records_per_second,
                )

                self.metadata_repository.update_sheet_source(
                    sheet_model,
                    row_count=row_count,
                    column_schema=self._build_column_schema(columns=columns, rows=rows),
                    checksum=sheet_checksum,
                    last_refreshed_at=completed_at,
                    position_index=position_index,
                )

                sheet_results.append(
                    SheetIngestionResult(
                        sheet=sheet_model,
                        data_file=data_file,
                        embedding_summary=embedding_summary,
                        processed_rows=processed_rows,
                        skipped_rows=skipped_rows,
                    )
                )

            self.metadata_repository.update_source_bundle(
                bundle,
                ingestion_status=IngestionStatus.READY,
                sheet_count=len(sheet_results),
            )

            summary = build_sheet_summary(included, hidden_opt_ins, inactive_sheets)
            audit = self.metadata_repository.create_bundle_audit(
                bundle=bundle,
                status=AuditStatus.SUCCEEDED,
                started_at=ingestion_started,
                completed_at=datetime.now(timezone.utc),
                sheet_summary=summary,
                hidden_sheets_enabled=[sheet.name for sheet in hidden_opt_ins],
            )

            self.metadata_repository.session.commit()  # type: ignore[attr-defined]

            return BundleIngestionResult(
                bundle=bundle,
                sheets=sheet_results,
                hidden_opt_ins=list(hidden_opt_ins),
                audit=audit,
            )
        except Exception as error:
            LOGGER.exception("Bundle ingestion failed for %s: %s", display_name, error)
            self.metadata_repository.update_source_bundle(
                bundle,
                ingestion_status=IngestionStatus.FAILED,
            )
            self.metadata_repository.create_bundle_audit(
                bundle=bundle,
                status=AuditStatus.FAILED,
                started_at=ingestion_started,
                completed_at=datetime.now(timezone.utc),
                sheet_summary=None,
                hidden_sheets_enabled=[sheet.name for sheet in hidden_opt_ins],
            )
            self.metadata_repository.session.commit()  # type: ignore[attr-defined]
            raise

    def refresh_bundle(
        self,
        *,
        bundle_id: str,
        hidden_sheet_policy: HiddenSheetPolicy,
        rename_tolerance: str = "allow_same_schema",
        source_path: Path | None = None,
    ) -> BundleRefreshResult:
        bundle = self.metadata_repository.get_source_bundle(bundle_id)
        if bundle is None:
            raise ValueError(f"Source bundle '{bundle_id}' not found.")

        ingestion_started = datetime.now(timezone.utc)

        if source_path is not None:
            stored_path = self._copy_bundle_file(source_path, bundle)
            self.metadata_repository.update_source_bundle(bundle, original_path=str(stored_path))
        else:
            if not bundle.original_path:
                raise ValueError("Bundle has no stored workbook path to refresh from.")
            stored_path = Path(bundle.original_path)
            if not stored_path.exists():
                raise FileNotFoundError(f"Stored bundle file missing: {stored_path}")

        file_type = bundle.file_type
        if file_type == FileType.EXCEL:
            discovered_sheets = discover_workbook_sheets(stored_path)
        else:
            discovered_sheets = [DiscoveredSheet(name="__csv__", position=0, hidden=False)]

        included, hidden_opt_ins, _excluded = apply_hidden_sheet_policy(discovered_sheets, hidden_sheet_policy)
        if not included:
            raise ValueError("No sheets available after applying hidden sheet policy.")

        existing_sheets = list(self.metadata_repository.list_sheet_sources(bundle_id=bundle.id))
        existing_map = {sheet.id: sheet for sheet in existing_sheets}
        existing_snapshots = [
            ExistingSheetSnapshot(
                id=sheet.id,
                sheet_name=sheet.sheet_name,
                checksum=sheet.checksum,
                column_schema=sheet.column_schema,
                row_count=sheet.row_count,
            )
            for sheet in existing_sheets
        ]

        discovered_payload: dict[str, dict[str, object]] = {}
        discovered_snapshots: list[DiscoveredSheetSnapshot] = []

        for sheet in included:
            sheet_name = "__csv__" if file_type == FileType.CSV else sheet.name
            visibility_state = (
                SheetVisibilityState.HIDDEN_OPT_IN if sheet.hidden else SheetVisibilityState.VISIBLE
            )
            ingestion_options = IngestionOptions(
                selected_columns=[],
                delimiter=bundle.delimiter,
                sheet_name=None if file_type == FileType.CSV else sheet_name,
                encoding="utf-8",
            )
            columns, rows = self._load_rows(stored_path, file_type, ingestion_options)
            selected_columns = [column for column in columns if column]
            column_schema = self._build_column_schema(columns=columns, rows=rows)
            checksum = self._hash_sheet_content(columns=columns, rows=rows)

            discovered_snapshots.append(
                DiscoveredSheetSnapshot(
                    sheet_name=sheet_name,
                    checksum=checksum,
                    column_schema=column_schema,
                    row_count=len(rows),
                )
            )
            discovered_payload[sheet_name] = {
                "sheet_name": sheet_name,
                "position": sheet.position,
                "visibility_state": visibility_state,
                "columns": columns,
                "rows": rows,
                "selected_columns": selected_columns,
                "column_schema": column_schema,
                "checksum": checksum,
                "row_count": len(rows),
            }

        matches, unmatched_existing = match_sheets(
            existing_snapshots,
            discovered_snapshots,
            tolerance=rename_tolerance,
        )

        created_results: list[SheetIngestionResult] = []
        updated_results: list[SheetIngestionResult] = []
        deactivated_sheets: list[SheetSource] = []

        try:
            for discovered_snapshot, existing_snapshot in matches:
                payload = discovered_payload[discovered_snapshot.sheet_name]
                row_count = int(payload["row_count"])
                columns = list(payload["columns"])
                rows = list(payload["rows"])
                selected_columns = list(payload["selected_columns"])
                column_schema = list(payload["column_schema"])
                visibility_state = payload["visibility_state"]
                position_index = int(payload["position"])
                checksum = str(payload["checksum"])

                if existing_snapshot is None:
                    if row_count == 0:
                        continue
                    sheet_model = self.metadata_repository.create_sheet_source(
                        bundle=bundle,
                        sheet_name=discovered_snapshot.sheet_name,
                        display_label=self._build_display_label(bundle, file_type, discovered_snapshot.sheet_name),
                        visibility_state=visibility_state,
                        status=SheetStatus.ACTIVE,
                        row_count=0,
                        column_schema=[],
                        position_index=position_index,
                        checksum=None,
                    )
                    self.metadata_repository.session.flush()  # type: ignore[attr-defined]
                    try:
                        result = self._ingest_refreshed_sheet(
                            bundle=bundle,
                            sheet_model=sheet_model,
                            stored_path=stored_path,
                            file_type=file_type,
                            sheet_name=discovered_snapshot.sheet_name,
                            columns=columns,
                            rows=rows,
                            selected_columns=selected_columns,
                            column_schema=column_schema,
                            checksum=checksum,
                            position_index=position_index,
                            visibility_state=visibility_state,
                        )
                    except ValueError as error:
                        LOGGER.warning(
                            "Skipping new sheet '%s' during refresh due to missing columns: %s",
                            discovered_snapshot.sheet_name,
                            error,
                        )
                        now = datetime.now(timezone.utc)
                        self.metadata_repository.update_sheet_source(
                            sheet_model,
                            status=SheetStatus.INACTIVE,
                            row_count=row_count,
                            column_schema=column_schema,
                            checksum=checksum,
                            last_refreshed_at=now,
                            position_index=position_index,
                        )
                        data_file = self.metadata_repository.get_data_file_by_hash(
                            self._hash_sheet_identifier(bundle.file_hash, discovered_snapshot.sheet_name)
                        )
                        if data_file is not None:
                            self.metadata_repository.update_data_file_status(
                                data_file,
                                status=IngestionStatus.FAILED,
                                row_count=row_count,
                                error_summary=str(error),
                                processed_at=now,
                            )
                            self.metadata_repository.create_audit(
                                data_file_id=data_file.id,
                                status=AuditStatus.FAILED,
                                processed_rows=0,
                                skipped_rows=0,
                                started_at=now,
                                completed_at=now,
                            )
                        deactivated_sheets.append(sheet_model)
                        continue
                    created_results.append(result)
                    continue

                sheet_model = existing_map[existing_snapshot.id]

                if row_count == 0:
                    self.metadata_repository.update_sheet_source(
                        sheet_model,
                        status=SheetStatus.INACTIVE,
                        row_count=0,
                        column_schema=[],
                        checksum=None,
                        last_refreshed_at=datetime.now(timezone.utc),
                        position_index=position_index,
                    )
                    deactivated_sheets.append(sheet_model)
                    continue

                try:
                    result = self._ingest_refreshed_sheet(
                        bundle=bundle,
                        sheet_model=sheet_model,
                        stored_path=stored_path,
                        file_type=file_type,
                        sheet_name=discovered_snapshot.sheet_name,
                        columns=columns,
                        rows=rows,
                        selected_columns=selected_columns,
                        column_schema=column_schema,
                        checksum=checksum,
                        position_index=position_index,
                        visibility_state=visibility_state,
                    )
                except ValueError as error:
                    LOGGER.warning(
                        "Marking sheet '%s' inactive during refresh due to missing columns: %s",
                        discovered_snapshot.sheet_name,
                        error,
                    )
                    now = datetime.now(timezone.utc)
                    self.metadata_repository.update_sheet_source(
                        sheet_model,
                        status=SheetStatus.INACTIVE,
                        row_count=row_count,
                        column_schema=column_schema,
                        checksum=checksum,
                        last_refreshed_at=now,
                        position_index=position_index,
                    )
                    data_file = self.metadata_repository.get_data_file_by_hash(
                        self._hash_sheet_identifier(bundle.file_hash, discovered_snapshot.sheet_name)
                    )
                    if data_file is not None:
                        self.metadata_repository.update_data_file_status(
                            data_file,
                            status=IngestionStatus.FAILED,
                            row_count=row_count,
                            error_summary=str(error),
                            processed_at=now,
                        )
                        self.metadata_repository.create_audit(
                            data_file_id=data_file.id,
                            status=AuditStatus.FAILED,
                            processed_rows=0,
                            skipped_rows=0,
                            started_at=now,
                            completed_at=now,
                        )
                    deactivated_sheets.append(sheet_model)
                    continue
                updated_results.append(result)

            for snapshot in unmatched_existing:
                sheet_model = existing_map[snapshot.id]
                self.metadata_repository.update_sheet_source(
                    sheet_model,
                    status=SheetStatus.INACTIVE,
                    last_refreshed_at=datetime.now(timezone.utc),
                )
                deactivated_sheets.append(sheet_model)

            active_count = sum(1 for sheet in self.metadata_repository.list_sheet_sources(bundle_id=bundle.id) if sheet.status == SheetStatus.ACTIVE)
            self.metadata_repository.update_source_bundle(
                bundle,
                ingestion_status=IngestionStatus.READY,
                sheet_count=active_count,
            )

            audit = self.metadata_repository.create_bundle_audit(
                bundle=bundle,
                status=AuditStatus.SUCCEEDED,
                started_at=ingestion_started,
                completed_at=datetime.now(timezone.utc),
                sheet_summary={
                    "created": len(created_results),
                    "updated": len(updated_results),
                    "deactivated": len(deactivated_sheets),
                },
                hidden_sheets_enabled=[sheet.name for sheet in hidden_opt_ins],
            )

            self.metadata_repository.session.commit()  # type: ignore[attr-defined]

            return BundleRefreshResult(
                bundle=bundle,
                created=created_results,
                updated=updated_results,
                deactivated=deactivated_sheets,
                audit=audit,
            )
        except Exception as error:
            LOGGER.exception("Bundle refresh failed for %s: %s", bundle.display_name, error)
            self.metadata_repository.update_source_bundle(
                bundle,
                ingestion_status=IngestionStatus.PARTIAL_FAILED,
            )
            self.metadata_repository.create_bundle_audit(
                bundle=bundle,
                status=AuditStatus.FAILED,
                started_at=ingestion_started,
                completed_at=datetime.now(timezone.utc),
                sheet_summary=None,
                hidden_sheets_enabled=[sheet.name for sheet in hidden_opt_ins],
            )
            self.metadata_repository.session.commit()  # type: ignore[attr-defined]
            raise

    def _hash_file(self, path: Path) -> str:
        sha = hashlib.sha256()
        with path.open("rb") as handle:
            for chunk in iter(lambda: handle.read(8192), b""):
                sha.update(chunk)
        return sha.hexdigest()

    def _ingest_refreshed_sheet(
        self,
        *,
        bundle: SourceBundle,
        sheet_model: SheetSource,
        stored_path: Path,
        file_type: FileType,
        sheet_name: str,
        columns: Sequence[str],
        rows: list[dict[str, object]],
        selected_columns: Sequence[str],
        column_schema: Sequence[dict[str, object]],
        checksum: str,
        position_index: int,
        visibility_state: SheetVisibilityState,
    ) -> SheetIngestionResult:
        sheet_timer = time.perf_counter()
        sheet_started = datetime.now(timezone.utc)

        self.metadata_repository.delete_query_records_for_sheet(sheet_model.id)

        sheet_hash = self._hash_sheet_identifier(bundle.file_hash, sheet_name)
        display_label = self._build_display_label(bundle, file_type, sheet_name)
        selected = list(selected_columns) or [column for column in columns if column]

        data_file = self.metadata_repository.get_data_file_by_hash(sheet_hash)
        if data_file is None:
            data_file = self.metadata_repository.create_data_file(
                display_name=display_label,
                original_path=str(stored_path),
                file_hash=sheet_hash,
                file_type=file_type,
                delimiter=bundle.delimiter if file_type == FileType.CSV else None,
                sheet_name=sheet_name if file_type == FileType.EXCEL else "__csv__",
                selected_columns=selected,
                status=IngestionStatus.PROCESSING,
            )
            self.metadata_repository.session.flush()  # type: ignore[attr-defined]
        else:
            data_file.display_name = display_label
            data_file.original_path = str(stored_path)
            data_file.file_type = file_type
            data_file.delimiter = bundle.delimiter if file_type == FileType.CSV else None
            data_file.sheet_name = sheet_name if file_type == FileType.EXCEL else "__csv__"
            data_file.selected_columns = list(selected)
            self.metadata_repository.update_data_file_status(
                data_file,
                status=IngestionStatus.PROCESSING,
                row_count=0,
                error_summary=None,
                processed_at=None,
            )

        (
            records,
            skipped_rows,
            available_columns,
            missing_columns,
        ) = self._materialize_records(
            rows=rows,
            columns=columns,
            data_file=data_file,
            selected_columns=selected,
            sheet_id=sheet_model.id,
        )
        if not available_columns:
            raise ValueError(f"Selected columns not found for sheet '{sheet_name}'.")

        data_file.selected_columns = list(available_columns)
        if missing_columns:
            LOGGER.warning(
                "Refresh detected missing columns on sheet '%s'; continuing with available subset: %s",
                sheet_name,
                missing_columns,
            )

        processed_rows = len(records)
        if processed_rows == 0:
            raise ValueError(f"No valid textual rows found for sheet '{sheet_name}'.")

        embedding_summary = self.embedding_service.run_embedding(
            EmbeddingJob(
                data_file=data_file,
                sheet=sheet_model,
                records=records,
                metadata_repository=self.metadata_repository,
            )
        )

        completed_at = datetime.now(timezone.utc)
        self.metadata_repository.update_data_file_status(
            data_file,
            status=IngestionStatus.READY,
            row_count=len(rows),
            error_summary=None,
            processed_at=completed_at,
        )

        elapsed_ms = (time.perf_counter() - sheet_timer) * 1000.0
        records_per_second = processed_rows / (elapsed_ms / 1000.0) if elapsed_ms else None

        self.metadata_repository.create_audit(
            data_file_id=data_file.id,
            status=AuditStatus.SUCCEEDED,
            processed_rows=processed_rows,
            skipped_rows=skipped_rows,
            started_at=sheet_started,
            completed_at=completed_at,
        )

        self.metadata_repository.record_performance_metric(
            metric_type=MetricType.INGESTION,
            data_file_id=data_file.id,
            cluster_id=None,
            benchmark_run_id=None,
            p50_ms=elapsed_ms,
            p95_ms=elapsed_ms,
            records_per_second=records_per_second,
        )

        self.metadata_repository.update_sheet_source(
            sheet_model,
            sheet_name=sheet_name,
            display_label=display_label,
            visibility_state=visibility_state,
            status=SheetStatus.ACTIVE,
            row_count=len(rows),
            column_schema=column_schema,
            checksum=checksum,
            last_refreshed_at=completed_at,
            position_index=position_index,
        )

        return SheetIngestionResult(
            sheet=sheet_model,
            data_file=data_file,
            embedding_summary=embedding_summary,
            processed_rows=processed_rows,
            skipped_rows=skipped_rows,
        )

    def _build_display_label(self, bundle: SourceBundle, file_type: FileType, sheet_name: str) -> str:
        if file_type == FileType.CSV:
            return bundle.display_name
        return f"{bundle.display_name}:{sheet_name}"

    def _infer_file_type(self, path: Path) -> FileType:
        extension = path.suffix.lower()
        if extension in {".csv", ".tsv"}:
            return FileType.CSV
        if extension in {".xls", ".xlsx"}:
            return FileType.EXCEL
        raise ValueError(f"Unsupported file type: {extension}")

    def _copy_raw_file(self, source: Path, data_file: DataFile) -> Path:
        destination_dir = self.data_root / "raw" / data_file.id
        destination_dir.mkdir(parents=True, exist_ok=True)
        destination = destination_dir / source.name
        shutil.copy2(source, destination)
        return destination

    def _copy_bundle_file(self, source: Path, bundle: SourceBundle) -> Path:
        destination_dir = self.data_root / "bundles" / bundle.id
        destination_dir.mkdir(parents=True, exist_ok=True)
        destination = destination_dir / source.name
        shutil.copy2(source, destination)
        return destination

    def _load_rows(
        self,
        path: Path,
        file_type: FileType,
        options: IngestionOptions,
    ) -> tuple[list[str], list[dict[str, object]]]:
        if file_type == FileType.CSV:
            delimiter = options.delimiter or ","
            with path.open("r", encoding=options.encoding, newline="") as handle:
                reader = csv.DictReader(handle, delimiter=delimiter)
                rows = [dict(row) for row in reader]
                columns = reader.fieldnames or []
            return columns, rows

        if load_workbook is None:
            raise RuntimeError("Excel ingestion requires the 'openpyxl' dependency.")

        workbook = load_workbook(path, read_only=True, data_only=True)
        sheet_name = options.sheet_name or workbook.sheetnames[0]
        if sheet_name not in workbook.sheetnames:
            raise ValueError(f"Sheet '{sheet_name}' not found in workbook.")

        sheet = workbook[sheet_name]
        iterator = sheet.iter_rows(values_only=True)
        try:
            headers_row = next(iterator)
        except StopIteration:
            return [], []

        columns = [str(value) if value is not None else "" for value in headers_row]
        rows: list[dict[str, object]] = []
        for values in iterator:
            row_map: dict[str, object] = {}
            for index, column in enumerate(columns):
                if not column:
                    continue
                cell_value = None
                if values is not None and index < len(values):
                    cell_value = values[index]
                row_map[column] = cell_value
            rows.append(row_map)
        workbook.close()
        return columns, rows

    def _materialize_records(
        self,
        *,
        rows: list[dict[str, object]],
        columns: Sequence[str],
        data_file: DataFile,
        selected_columns: Sequence[str],
        sheet_id: str | None = None,
    ) -> tuple[list[QueryRecord], int, list[str], list[str]]:
        missing = [column for column in selected_columns if column not in columns]
        available = [column for column in selected_columns if column in columns]
        if missing:
            LOGGER.warning(
                "Missing columns during ingestion for data_file %s: %s",
                data_file.id,
                missing,
            )
        records: list[QueryRecord] = []
        skipped = 0

        for index, row in enumerate(rows):
            for column in available:
                original_value = row.get(column)
                text = self._normalize_text(original_value)
                if not text:
                    skipped += 1
                    continue
                record = self.metadata_repository.create_query_record(
                    data_file_id=data_file.id,
                    column_name=str(column),
                    row_index=int(index),
                    text=text,
                    original_text=str(original_value) if original_value is not None else "",
                    tags=None,
                    sheet_id=sheet_id,
                )
                records.append(record)
        self.metadata_repository.session.flush()  # type: ignore[attr-defined]
        return records, skipped, available, missing

    def _normalize_text(self, value: object) -> Optional[str]:
        if value is None:
            return None
        text = str(value).strip()
        if not text:
            return None
        if text.isnumeric():
            return None
        text = " ".join(text.split())
        return text

    def _build_column_schema(
        self,
        *,
        columns: Sequence[str],
        rows: Sequence[dict[str, object]],
    ) -> list[dict[str, object]]:
        schema: list[dict[str, object]] = []
        for column in columns:
            if not column:
                continue
            values = [row.get(column) for row in rows if column in row]
            inferred_type = self._infer_value_type(values)
            nullable = any(value is None for value in values) or not values
            schema.append(
                {
                    "name": column,
                    "inferredType": inferred_type,
                    "nullable": nullable,
                }
            )
        return schema

    def _infer_value_type(self, values: Sequence[object | None]) -> str:
        non_null = [value for value in values if value is not None]
        if not non_null:
            return "string"
        if all(isinstance(value, (int, float)) for value in non_null):
            return "number"
        if all(isinstance(value, bool) for value in non_null):
            return "boolean"
        return "string"

    def _hash_sheet_identifier(self, bundle_hash: str, sheet_name: str) -> str:
        digest = hashlib.sha256()
        digest.update(bundle_hash.encode("utf-8"))
        digest.update(b":")
        digest.update(sheet_name.encode("utf-8"))
        return digest.hexdigest()

    def _hash_sheet_content(
        self,
        *,
        columns: Sequence[str],
        rows: Sequence[dict[str, object]],
    ) -> str:
        payload = json.dumps(
            {
                "columns": list(columns),
                "rows": rows,
            },
            sort_keys=True,
            default=str,
        )
        digest = hashlib.sha256()
        digest.update(payload.encode("utf-8"))
        return digest.hexdigest()
