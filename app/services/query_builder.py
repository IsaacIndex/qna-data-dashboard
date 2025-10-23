from __future__ import annotations

import csv
import hashlib
import math
import time
from dataclasses import dataclass, field
from pathlib import Path
from typing import Sequence

try:  # Optional dependency: required for Excel preview support.
    from openpyxl import load_workbook
except ImportError:  # pragma: no cover - optional path in production
    load_workbook = None  # type: ignore[assignment]

from app.db.metadata import MetadataRepository
from app.db.schema import FileType, QuerySheetRole, SheetSource, SheetStatus


class QueryValidationError(RuntimeError):
    """Raised when a preview request fails validation."""


class QueryConflictError(QueryValidationError):
    """Raised when a query cannot execute due to sheet state or schema conflict."""


@dataclass(slots=True)
class QuerySheetSelection:
    sheet_id: str
    alias: str
    role: QuerySheetRole = QuerySheetRole.PRIMARY
    join_keys: Sequence[str] = field(default_factory=tuple)


@dataclass(slots=True)
class QueryProjection:
    expression: str
    label: str


@dataclass(slots=True)
class QueryFilter:
    sheet_alias: str
    column: str
    operator: str
    value: object


@dataclass(slots=True)
class QueryPreviewRequest:
    sheets: Sequence[QuerySheetSelection]
    projections: Sequence[QueryProjection]
    filters: Sequence[QueryFilter] = field(default_factory=tuple)
    limit: int | None = None


@dataclass(slots=True)
class QueryPreviewResult:
    headers: list[str]
    rows: list[list[str]]
    warnings: list[str]
    execution_ms: float
    row_count: int


class QueryBuilderService:
    """Execute sheet-level preview queries with lightweight validation."""

    def __init__(self, *, metadata_repository: MetadataRepository) -> None:
        self.metadata_repository = metadata_repository

    def preview_query(self, request: QueryPreviewRequest) -> QueryPreviewResult:
        if not request.sheets:
            raise QueryValidationError("At least one sheet must be selected.")
        if not request.projections:
            raise QueryValidationError("At least one projection must be provided.")

        start = time.perf_counter()

        alias_map: dict[str, QuerySheetSelection] = {}
        sheet_map: dict[str, SheetSource] = {}
        warnings: list[str] = []

        primary_selection: QuerySheetSelection | None = None
        primary_sheet: SheetSource | None = None
        for selection in request.sheets:
            if selection.alias in alias_map:
                raise QueryValidationError(f"Duplicate sheet alias '{selection.alias}'.")

            sheet = self.metadata_repository.get_sheet_source(selection.sheet_id)
            if sheet is None:
                raise QueryValidationError(f"Sheet '{selection.sheet_id}' not found.")
            if sheet.status != SheetStatus.ACTIVE:
                warnings.append(
                    f"Sheet '{selection.alias}' ({sheet.display_label}) is {sheet.status.value}."
                )

            if primary_selection is None or selection.role == QuerySheetRole.PRIMARY:
                primary_selection = selection
                primary_sheet = sheet

            alias_map[selection.alias] = selection
            sheet_map[selection.alias] = sheet

        assert primary_selection is not None  # for type checking
        assert primary_sheet is not None

        # Load primary sheet rows to seed the result set.
        primary_rows = self._load_sheet_rows(primary_sheet)
        combined_rows: list[dict[str, dict[str, object]]] = [
            {primary_selection.alias: row} for row in primary_rows
        ]
        # Join additional sheets sequentially.
        for selection in request.sheets:
            if selection.alias == primary_selection.alias:
                continue
            if selection.role == QuerySheetRole.UNION:
                raise QueryValidationError("Union operations are not supported in preview mode.")
            if selection.role == QuerySheetRole.JOIN and not selection.join_keys:
                raise QueryValidationError(f"Join keys required for alias '{selection.alias}'.")

            join_sheet = sheet_map[selection.alias]
            join_rows = self._load_sheet_rows(join_sheet)
            join_warnings = self.validate_join_keys(
                primary_sheet.column_schema,
                join_sheet.column_schema,
                join_keys=selection.join_keys,
                primary_alias=primary_selection.alias,
                join_alias=selection.alias,
            )
            warnings.extend(join_warnings)
            combined_rows = self._join_rows(
                combined_rows=combined_rows,
                primary_alias=primary_selection.alias,
                join_alias=selection.alias,
                join_rows=join_rows,
                join_keys=selection.join_keys,
            )

        filtered_rows = self._apply_filters(
            combined_rows=combined_rows,
            filters=request.filters,
            alias_map=alias_map,
        )
        projected = self._project_rows(
            rows=filtered_rows,
            projections=request.projections,
            alias_map=alias_map,
            primary_alias=primary_selection.alias,
        )

        limited_rows = projected
        if request.limit is not None:
            limited_rows = projected[: max(0, request.limit)]

        execution_ms = (time.perf_counter() - start) * 1000.0

        return QueryPreviewResult(
            headers=[projection.label for projection in request.projections],
            rows=limited_rows,
            warnings=warnings,
            execution_ms=execution_ms,
            row_count=len(limited_rows),
        )

    @staticmethod
    def validate_join_keys(
        primary_schema: Sequence[dict[str, object]],
        join_schema: Sequence[dict[str, object]],
        *,
        join_keys: Sequence[str],
        primary_alias: str,
        join_alias: str,
    ) -> list[str]:
        if not join_keys:
            raise QueryValidationError(f"Join keys required for alias '{join_alias}'.")

        primary_lookup = {
            str(column.get("name")): column for column in primary_schema if column.get("name") is not None
        }
        join_lookup = {str(column.get("name")): column for column in join_schema if column.get("name") is not None}

        warnings: list[str] = []
        for key in join_keys:
            if key not in primary_lookup:
                raise QueryValidationError(
                    f"Join column '{key}' missing on sheet alias '{primary_alias}'."
                )
            if key not in join_lookup:
                raise QueryValidationError(
                    f"Join column '{key}' missing on sheet alias '{join_alias}'."
                )

            primary_type = str(primary_lookup[key].get("inferredType") or "").lower()
            join_type = str(join_lookup[key].get("inferredType") or "").lower()
            if primary_type and join_type and primary_type != join_type:
                warnings.append(
                    f"Join column '{key}' uses incompatible types between "
                    f"'{primary_alias}' ({primary_type}) and '{join_alias}' ({join_type})."
                )
        return warnings

    def _load_sheet_rows(self, sheet: SheetSource) -> list[dict[str, object]]:
        bundle = sheet.bundle
        sheet_hash = self._hash_sheet_identifier(bundle.file_hash, sheet.sheet_name)
        data_file = self.metadata_repository.get_data_file_by_hash(sheet_hash)
        if data_file is None:
            raise QueryValidationError(
                f"Backing data file for sheet '{sheet.display_label}' not found."
            )

        path = Path(data_file.original_path)
        if not path.exists():
            raise QueryValidationError(f"Sheet data file missing on disk: {path}")

        if data_file.file_type == FileType.CSV:
            return self._load_csv_rows(path=path, delimiter=data_file.delimiter or ",")

        if data_file.file_type == FileType.EXCEL:
            if load_workbook is None:
                raise QueryValidationError("Excel preview requires the 'openpyxl' dependency.")
            return self._load_excel_rows(path=path, sheet_name=sheet.sheet_name)

        raise QueryValidationError(f"Unsupported file type for preview: {data_file.file_type.value}")

    def _load_csv_rows(self, *, path: Path, delimiter: str) -> list[dict[str, object]]:
        with path.open("r", encoding="utf-8", newline="") as handle:
            reader = csv.DictReader(handle, delimiter=delimiter)
            return [dict(row) for row in reader]

    def _load_excel_rows(self, *, path: Path, sheet_name: str) -> list[dict[str, object]]:
        if load_workbook is None:  # pragma: no cover - guarded earlier
            raise QueryValidationError("Excel preview requires the 'openpyxl' dependency.")

        workbook = load_workbook(path, read_only=True, data_only=True)
        try:
            if sheet_name not in workbook.sheetnames:
                raise QueryValidationError(f"Sheet '{sheet_name}' not found in workbook.")

            worksheet = workbook[sheet_name]
            iterator = worksheet.iter_rows(values_only=True)
            try:
                headers_row = next(iterator)
            except StopIteration:
                return []

            headers = [str(value) if value is not None else "" for value in headers_row]
            rows: list[dict[str, object]] = []
            for values in iterator:
                row_map: dict[str, object] = {}
                for idx, header in enumerate(headers):
                    if not header:
                        continue
                    cell = None
                    if values is not None and idx < len(values):
                        cell = values[idx]
                    row_map[header] = cell
                rows.append(row_map)
            return rows
        finally:
            workbook.close()

    def _join_rows(
        self,
        *,
        combined_rows: list[dict[str, dict[str, object]]],
        primary_alias: str,
        join_alias: str,
        join_rows: list[dict[str, object]],
        join_keys: Sequence[str],
    ) -> list[dict[str, dict[str, object]]]:
        if not combined_rows:
            return []

        index: dict[tuple[object, ...], list[dict[str, object]]] = {}
        for row in join_rows:
            key = tuple(row.get(key) for key in join_keys)
            index.setdefault(key, []).append(row)

        result: list[dict[str, dict[str, object]]] = []
        for merged in combined_rows:
            primary_row = merged.get(primary_alias)
            if primary_row is None:
                raise QueryValidationError(f"Primary alias '{primary_alias}' missing in join context.")
            key = tuple(primary_row.get(join_key) for join_key in join_keys)
            matches = index.get(key)
            if not matches:
                continue
            for match in matches:
                combined = dict(merged)
                combined[join_alias] = match
                result.append(combined)
        return result

    def _apply_filters(
        self,
        *,
        combined_rows: list[dict[str, dict[str, object]]],
        filters: Sequence[QueryFilter],
        alias_map: dict[str, QuerySheetSelection],
    ) -> list[dict[str, dict[str, object]]]:
        rows = combined_rows
        for query_filter in filters:
            alias = query_filter.sheet_alias
            if alias not in alias_map:
                raise QueryValidationError(f"Filter references unknown alias '{alias}'.")
            operator = query_filter.operator.lower()

            filtered: list[dict[str, dict[str, object]]] = []
            for merged in rows:
                sheet_row = merged.get(alias)
                if sheet_row is None:
                    continue
                value = sheet_row.get(query_filter.column)
                if self._matches_filter(value, operator, query_filter.value):
                    filtered.append(merged)
            rows = filtered
        return rows

    def _matches_filter(self, value: object, operator: str, expected: object) -> bool:
        if operator == "eq":
            return value == expected
        if operator == "ne":
            return value != expected
        if operator == "contains":
            return isinstance(value, str) and isinstance(expected, str) and expected.lower() in value.lower()
        if operator in {"gt", "lt"}:
            lhs = self._coerce_number(value)
            rhs = self._coerce_number(expected)
            if lhs is None or rhs is None:
                return False
            if operator == "gt":
                return lhs > rhs
            return lhs < rhs
        raise QueryValidationError(f"Unsupported filter operator '{operator}'.")

    def _project_rows(
        self,
        *,
        rows: list[dict[str, dict[str, object]]],
        projections: Sequence[QueryProjection],
        alias_map: dict[str, QuerySheetSelection],
        primary_alias: str,
    ) -> list[list[str]]:
        aggregate_projections: list[tuple[str, str, QueryProjection]] = []
        scalar_projections: list[tuple[str, str, QueryProjection]] = []
        for projection in projections:
            parsed = self._parse_aggregate(projection.expression)
            if parsed is not None:
                aggregate_projections.append((*parsed, projection))
            else:
                alias, column = self._parse_column_reference(
                    projection.expression, alias_map=alias_map, primary_alias=primary_alias
                )
                scalar_projections.append((alias, column, projection))

        if aggregate_projections and scalar_projections:
            raise QueryValidationError("Cannot mix aggregate and scalar projections in preview.")

        if not rows and not aggregate_projections:
            return []

        if aggregate_projections:
            return [
                [
                    self._stringify_value(
                        self._execute_aggregate(
                            func=func,
                            alias=alias,
                            column=column,
                            rows=rows,
                        )
                    )
                    for func, alias, column, _projection in aggregate_projections
                ]
            ]

        projected_rows: list[list[str]] = []
        for merged in rows:
            projected_row: list[str] = []
            for alias, column, _projection in scalar_projections:
                sheet_row = merged.get(alias)
                if sheet_row is None:
                    projected_row.append("")
                    continue
                projected_row.append(self._stringify_value(sheet_row.get(column)))
            projected_rows.append(projected_row)
        return projected_rows

    def _parse_aggregate(self, expression: str) -> tuple[str, str, str] | None:
        expr = expression.strip()
        if "(" not in expr or not expr.endswith(")"):
            return None
        func, remainder = expr.split("(", 1)
        func = func.strip().lower()
        if func not in {"sum", "avg", "count"}:
            return None
        inner = remainder[:-1].strip()
        if not inner:
            raise QueryValidationError(f"Aggregate expression '{expression}' is empty.")
        if func == "count" and inner == "*":
            return func, "*", "*"
        alias, column = self._resolve_expression_alias(inner)
        return func, alias, column

    def _parse_column_reference(
        self,
        expression: str,
        *,
        alias_map: dict[str, QuerySheetSelection],
        primary_alias: str,
    ) -> tuple[str, str]:
        alias, column = self._resolve_expression_alias(expression)
        resolved_alias = alias or primary_alias
        if resolved_alias not in alias_map:
            raise QueryValidationError(f"Unknown sheet alias '{resolved_alias}' in projection '{expression}'.")
        if not column:
            raise QueryValidationError(f"Column missing in projection '{expression}'.")
        return resolved_alias, column

    def _resolve_expression_alias(self, expression: str) -> tuple[str | None, str]:
        cleaned = expression.strip()
        if "." in cleaned:
            alias, column = cleaned.split(".", 1)
            return alias.strip(), column.strip()
        return None, cleaned

    def _execute_aggregate(
        self,
        *,
        func: str,
        alias: str,
        column: str,
        rows: list[dict[str, dict[str, object]]],
    ) -> object:
        if alias != "*" and all(alias not in merged for merged in rows):
            raise QueryValidationError(f"Unknown sheet alias '{alias}' in aggregate '{func}'.")

        if func == "count":
            if alias == "*" and column == "*":
                return len(rows)
            total = 0
            for merged in rows:
                sheet_row = merged.get(alias)
                if sheet_row is None:
                    continue
                if sheet_row.get(column) is not None:
                    total += 1
            return total

        values: list[float] = []
        for merged in rows:
            sheet_row = merged.get(alias)
            if sheet_row is None:
                continue
            value = sheet_row.get(column)
            number = self._coerce_number(value)
            if number is None:
                continue
            values.append(number)

        if func == "sum":
            return sum(values)
        if func == "avg":
            return sum(values) / len(values) if values else 0.0
        raise QueryValidationError(f"Unsupported aggregate '{func}'.")

    def _stringify_value(self, value: object) -> str:
        if value is None:
            return ""
        if isinstance(value, float):
            if math.isfinite(value) and value.is_integer():
                return str(int(value))
            return f"{value:.6f}".rstrip("0").rstrip(".")
        return str(value)

    def _coerce_number(self, value: object) -> float | None:
        if isinstance(value, (int, float)):
            return float(value)
        if isinstance(value, str):
            try:
                return float(value)
            except ValueError:
                return None
        return None

    def _hash_sheet_identifier(self, bundle_hash: str, sheet_name: str) -> str:
        combined = f"{bundle_hash}:{sheet_name}".encode("utf-8")
        digest = hashlib.sha256()
        digest.update(combined)
        return digest.hexdigest()
